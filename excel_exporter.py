"""
excel_exporter.py
─────────────────────────────────────────────────────────────
Generates two formatted Excel workbooks from pricing results:
  1. cloud_sizing.xlsx   — node distribution + architecture
  2. aws_pricing.xlsx    — 5-year cost forecast with inflation

Uses openpyxl with professional formatting per SKILL.md standards:
  Blue text  = hardcoded inputs
  Black text = formulas / calculations
  Green fill = section headers
─────────────────────────────────────────────────────────────
"""

from __future__ import annotations
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ── Style helpers ─────────────────────────────────────────────────────────

def _font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _num_fmt(cell, fmt):
    cell.number_format = fmt

HEADER_FILL  = _fill("1F4E79")   # dark navy
SECTION_FILL = _fill("2E75B6")   # medium blue
ALT_FILL     = _fill("EBF3FB")   # light blue alternate row
TOTAL_FILL   = _fill("FFF2CC")   # yellow for totals
GREEN_FILL   = _fill("E2EFDA")   # light green for positives

def _header_row(ws, row, values, fill=None):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = _font(bold=True, color="FFFFFF", size=10)
        c.fill      = fill or _fill("1F4E79")
        c.alignment = _center()
        c.border    = _border()

def _data_row(ws, row, values, alt=False, bold=False, num_cols=None, total=False):
    fill = TOTAL_FILL if total else (_fill("EBF3FB") if alt else _fill("FFFFFF"))
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = _font(bold=bold or total, size=10)
        c.fill      = fill
        c.alignment = _left() if col <= 2 else _center()
        c.border    = _border()
        if num_cols and col in num_cols and isinstance(val, (int, float)):
            c.number_format = '$#,##0.00'


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _format_instance_with_specs(instance_type: str) -> str:
    """Appends basic specs to the instance type string (e.g. r5.4xlarge (16 vCPU, 128 GB RAM))."""
    if not instance_type or instance_type == "—" or " " in instance_type:
        return instance_type
        
    try:
        from aws_pricer import INSTANCE_SIZE_TABLES
        if "." in instance_type and not instance_type.startswith("db.") and not instance_type.startswith("cache.") and not instance_type.startswith("s3."):
            family, size = instance_type.split(".", 1)
            for c, r_mem, s in INSTANCE_SIZE_TABLES.get(family, []):
                if s == size:
                    return f"{instance_type} ({c} vCPU, {r_mem}GB RAM)"
    except Exception:
        pass
        
    return instance_type


# Category label normalisation shared by AWS and GCP sizing sheets
_GCP_CAT_MAP: dict[str, str] = {"S3": "Cloud Storage", "S3 Storage": "Cloud Storage"}

# ── Sheet: Prod Sizing ───────────────────────────────────────────────────

def _build_prod_sizing_sheet(
    wb: Workbook,
    distribution: dict,
    pricing: dict,
    metrics: dict,
    customer: str,
):
    """
    Production infrastructure sizing sheet.
    Columns: Category | Service/Role | Nodes | Instance Type | vCPU/Node | RAM/Node | Storage/Node | Pricing Model | Reasoning
    No monetary values.
    """
    ws = wb.active
    ws.title = "Prod Sizing"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    NCOLS = 9
    col_end = get_column_letter(NCOLS)

    # Build a lookup: role_key → instance_type from priced roles
    itype_map = {}
    for r in (pricing or {}).get("priced_roles", []):
        rk = r.get("role_key") or ""
        it = r.get("instance_type") or ""
        if rk and it:
            itype_map[rk] = _format_instance_with_specs(it)

    # ── Title block ──────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{col_end}1")
    title = ws["A1"]
    title.value     = f"Production Infrastructure Sizing  |  {customer}  |  {datetime.today().strftime('%d %b %Y')}"
    title.font      = _font(bold=True, size=13, color="FFFFFF")
    title.fill      = _fill("1F4E79")
    title.alignment = _center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{col_end}2")
    sub = ws["A2"]
    ch_nodes   = distribution.get("clickhouse_nodes", {})
    ch_enabled = ch_nodes and ch_nodes.get("enabled")
    ch_total   = ch_nodes.get("summary", {}).get("total_nodes", 0) if ch_enabled else 0
    ch_note    = f"   |   ClickHouse Nodes: {ch_total}" if ch_enabled else ""
    sub.value     = (
        f"Worker Nodes: {distribution['summary']['total_worker_nodes']}"
        f"   |   DB Nodes: {distribution['summary']['total_db_nodes']}"
        f"   |   Confidence: {distribution['summary']['confidence'].title()}"
        f"{ch_note}"
    )
    sub.font      = _font(italic=True, size=10, color="FFFFFF")
    sub.fill      = _fill("2E75B6")
    sub.alignment = _center()
    ws.row_dimensions[2].height = 20

    # ── Headers ──────────────────────────────────────────────────────────
    headers = [
        "Category", "Service / Role", "Nodes",
        "Instance Type", "vCPU / Node", "RAM / Node (GB)",
        "Storage / Node (GB)", "Pricing Model", "Reasoning",
    ]
    _header_row(ws, 3, headers)
    ws.row_dimensions[3].height = 32

    # ── Data rows ─────────────────────────────────────────────────────────
    all_roles = (
        distribution.get("worker_nodes", [])
        + distribution.get("db_nodes", [])
        + distribution.get("fixed_roles", [])
    )

    row = 4
    prev_cat = None
    for i, r in enumerate(all_roles):
        cat = r.get("category", "—")
        if cat != prev_cat:
            ws.merge_cells(f"A{row}:{col_end}{row}")
            c = ws.cell(row=row, column=1, value=f"  {cat}")
            c.font = _font(bold=True, size=10, color="FFFFFF")
            c.fill = _fill("2E75B6")
            c.alignment = _left()
            c.border = _border()
            ws.row_dimensions[row].height = 20
            row += 1
            prev_cat = cat

        rk        = r.get("role_key", "")
        itype     = itype_map.get(rk, "—")
        vcpu      = r.get("vcpu_per_node", "—") or "—"
        ram       = r.get("ram_per_node",  "—") or "—"
        stor      = r.get("storage_per_node_gb", 0) or "—"
        pm        = r.get("pricing_model") or "On Demand"
        reasoning = r.get("reasoning", "—")

        vals = [cat, r.get("label", "—"), r.get("nodes", 0),
                itype, vcpu, ram, stor, pm, reasoning]
        _data_row(ws, row, vals, alt=(i % 2 == 0))
        ws.row_dimensions[row].height = 18
        row += 1

    # ── Sizing inputs footer ──────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:{col_end}{row}")
    c = ws.cell(row=row, column=1, value="  Sizing Inputs (from Excel Template)")
    c.font = _font(bold=True, size=10, color="FFFFFF")
    c.fill = _fill("1F4E79")
    c.alignment = _left()
    row += 1

    db_ram_key = (
        "oracle_ram_gb" if metrics.get("db_type") == "Oracle"
        else "sql_server_ram_gb" if metrics.get("db_type") == "SQL Server"
        else "postgres_ram_gb"
    )
    metric_pairs = [
        ("Total Worker Nodes",   metrics.get("total_workernodes", 0)),
        ("Total vCPUs (Worker)", metrics.get("total_vcpus_workernode", 0)),
        ("Total RAM GB (Worker)",metrics.get("total_memory_workernode_gb", 0)),
        (f"{metrics.get('db_type','PostgreSQL')} Total RAM (GB)", metrics.get(db_ram_key, 0)),
        ("Data Size (GB)",       metrics.get("data_size_gb", 0)),
        ("S3 Size (GB)",         metrics.get("s3_size_gb", 0)),
    ]
    for i, (label, val) in enumerate(metric_pairs):
        _data_row(ws, row, [label, val] + [""] * (NCOLS - 2), alt=(i % 2 == 0))
        ws.cell(row=row, column=2).number_format = "#,##0"
        row += 1

    _set_col_widths(ws, [24, 42, 8, 20, 12, 15, 18, 16, 48])


# ── Sheet: Environment Sizing (Pre-Prod / DR) ─────────────────────────────

def _build_env_sizing_sheet(
    wb: Workbook,
    env_roles: list,
    sheet_name: str,
    title: str,
    customer: str,
):
    """
    Infrastructure sizing sheet for a non-production environment.
    Uses the priced_roles list (which contains instance_type from the pricer).
    NO monetary columns.
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    NCOLS = 8
    col_end = get_column_letter(NCOLS)

    ws.merge_cells(f"A1:{col_end}1")
    t = ws["A1"]
    t.value     = f"{title}  |  {customer}  |  {datetime.today().strftime('%d %b %Y')}"
    t.font      = _font(bold=True, size=13, color="FFFFFF")
    t.fill      = _fill("1F4E79")
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Summary count row
    total_nodes = sum(r.get("nodes", 0) for r in env_roles if r.get("nodes"))
    ws.merge_cells(f"A2:{col_end}2")
    s = ws["A2"]
    s.value     = f"Total Infrastructure Roles: {len(env_roles)}   |   Total Node Count: {total_nodes}"
    s.font      = _font(italic=True, size=10, color="FFFFFF")
    s.fill      = _fill("2E75B6")
    s.alignment = _center()
    ws.row_dimensions[2].height = 18

    headers = [
        "Category", "Service / Role", "Nodes",
        "Instance Type", "vCPU / Node", "RAM / Node (GB)",
        "Storage / Node (GB)", "Note",
    ]
    _header_row(ws, 3, headers)
    ws.row_dimensions[3].height = 28

    row = 4
    prev_cat = None
    for i, r in enumerate(env_roles):
        if r.get("monthly_usd", 0) == 0 and r.get("nodes", 0) == 0:
            continue
        raw_cat = r.get("category", "Other")
        cat = _GCP_CAT_MAP.get(raw_cat, raw_cat)
        if cat != prev_cat:
            ws.merge_cells(f"A{row}:{col_end}{row}")
            c = ws.cell(row=row, column=1, value=f"  {cat}")
            c.font = _font(bold=True, size=10, color="FFFFFF")
            c.fill = _fill("2E75B6")
            c.alignment = _left()
            c.border = _border()
            ws.row_dimensions[row].height = 20
            row += 1
            prev_cat = cat

        vcpu = r.get("vcpu", "—") or "—"
        ram  = r.get("ram",  "—") or "—"
        itype = _format_instance_with_specs(r.get("instance_type", "—") or "—")
        # derive storage from note if possible
        stor = "—"
        note_txt = r.get("note", "—") or "—"

        vals = [
            cat, r.get("label", "—"), r.get("nodes", "—"),
            itype, vcpu, ram, stor, note_txt,
        ]
        _data_row(ws, row, vals, alt=(i % 2 == 0))
        ws.row_dimensions[row].height = 18
        row += 1

    _set_col_widths(ws, [24, 45, 8, 20, 12, 15, 18, 50])

# ── Sheet 1: Pricing Summary ──────────────────────────────────────────────

def _build_pricing_summary_sheet(wb: Workbook, pricing: dict, env_pricing: dict, customer: str):
    ws = wb.active
    ws.title = "Pricing Summary"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    INFLATION = pricing.get("inflation_rate", 0.04)
    years     = 5

    total_cols = 3 + years + 1
    col_letter = get_column_letter(total_cols)

    # ── Title ─────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{col_letter}1")
    t = ws["A1"]
    t.value     = f"Cloud Pricing Summary  |  {customer}  |  {datetime.today().strftime('%d %b %Y')}  |  Inflation: {INFLATION*100:.1f}%/yr"
    t.font      = _font(bold=True, size=13, color="FFFFFF")
    t.fill      = _fill("1F4E79")
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{col_letter}2")
    s = ws["A2"]
    # We will compute grand totals later, let's leave this empty or with basic info
    s.value     = f"High-level 5-Year Cost Projection (All Environments)"
    s.font      = _font(italic=True, size=10, color="FFFFFF")
    s.fill      = _fill("2E75B6")
    s.alignment = _center()
    ws.row_dimensions[2].height = 20

    # ── Column headers ────────────────────────────────────────────────────
    year_labels = [f"Year {y+1}" for y in range(years)]
    headers = ["Environment", "Description", "Base Monthly (USD)"] + year_labels + ["5-Year Total"]
    _header_row(ws, 3, headers)
    ws.row_dimensions[3].height = 32

    # ── Inflation multipliers row ─────────────────────────────────────────
    mult_vals = ["", "Inflation Multiplier", "1.00"] + [f"={(1+INFLATION)**(y+1):.4f}" for y in range(years)] + [""]
    for col, val in enumerate(mult_vals, 1):
        c = ws.cell(row=4, column=col, value=val if not val.startswith("=") else None)
        if val.startswith("="):
            c.value = float(val[1:])
        c.font      = _font(italic=True, size=9, color="666666")
        c.fill      = _fill("F2F2F2")
        c.alignment = _center()
        c.border    = _border()
        if col >= 3:
            c.number_format = "0.0000"
    ws.row_dimensions[4].height = 16

    num_cols = set(range(3, 3 + years + 1 + 1))
    row = 5

    # 1. Prod
    prod_base = pricing.get("total_monthly_usd", 0)
    prod_yrs  = [round(prod_base * 12 * ((1 + INFLATION) ** y), 2) for y in range(years)]
    prod_5yr  = round(sum(prod_yrs), 2)
    _data_row(ws, row, ["Production", "Core Application & DB", prod_base] + prod_yrs + [prod_5yr], alt=False, num_cols=num_cols)
    row += 1

    # 2. Pre-Prod
    preprod_base = 0
    preprod_yrs  = [0] * years
    preprod_5yr  = 0
    if env_pricing and env_pricing.get("preprod_sit_uat"):
        preprod = env_pricing["preprod_sit_uat"]
        pp_names = preprod.get("env_names", [])
        pp_label = " + ".join(pp_names) if pp_names else "Pre-Prod / SIT / UAT"
        
        preprod_base = preprod.get("monthly_usd", 0)
        preprod_yrs  = [round(preprod_base * 12 * ((1 + INFLATION) ** y), 2) for y in range(years)]
        preprod_5yr  = round(sum(preprod_yrs), 2)
        _data_row(ws, row, [pp_label, "Testing & Staging Environment", preprod_base] + preprod_yrs + [preprod_5yr], alt=True, num_cols=num_cols)
        row += 1

    # 3. DR
    dr_base = 0
    dr_yrs  = [0] * years
    dr_5yr  = 0
    if env_pricing and env_pricing.get("dr"):
        dr_base = env_pricing["dr"].get("monthly_usd", 0)
        # DR might have its own inflation calc but we apply the same here for summary
        dr_yrs  = [round(dr_base * 12 * ((1 + INFLATION) ** y), 2) for y in range(years)]
        dr_5yr  = round(sum(dr_yrs), 2)
        _data_row(ws, row, ["Disaster Recovery", "DR Environment", dr_base] + dr_yrs + [dr_5yr], alt=False, num_cols=num_cols)
        row += 1

    # Grand Total
    grand_base = prod_base + preprod_base + dr_base
    grand_yrs  = [prod_yrs[y] + preprod_yrs[y] + dr_yrs[y] for y in range(years)]
    grand_5yr  = prod_5yr + preprod_5yr + dr_5yr
    _data_row(ws, row, ["GRAND TOTAL (All Environments)", "", grand_base] + grand_yrs + [grand_5yr], total=True, num_cols=num_cols)
    ws.row_dimensions[row].height = 22
    row += 2

    # ── Category Subtotals (Prod only for now to keep it clean) ───────────
    ws.merge_cells(f"A{row}:{col_letter}{row}")
    c = ws.cell(row=row, column=1, value="  Production Category Subtotals")
    c.font = _font(bold=True, size=10, color="FFFFFF"); c.fill = _fill("1F4E79")
    c.alignment = _left(); c.border = _border()
    row += 1

    for i, (cat, base_monthly) in enumerate(sorted(pricing.get("category_totals", {}).items())):
        year_costs = [round(base_monthly * 12 * ((1 + INFLATION) ** y), 2) for y in range(years)]
        five_yr    = round(sum(year_costs), 2)
        vals = [cat, "All services", base_monthly] + year_costs + [five_yr]
        _data_row(ws, row, vals, alt=(i % 2 == 0), num_cols=num_cols)
        row += 1

    _set_col_widths(ws, [26, 40, 18] + [16] * years + [16])


# ── Sheet 2: Prod Pricing ──────────────────────────────────────────────────

def _build_prod_pricing_sheet(wb: Workbook, pricing: dict, customer: str):
    ws = wb.create_sheet("Prod Pricing")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    INFLATION = pricing.get("inflation_rate", 0.04)
    years     = 5

    total_cols = 5 + years + 1
    col_letter = get_column_letter(total_cols)

    ws.merge_cells(f"A1:{col_letter}1")
    t = ws["A1"]
    t.value     = f"Production Environment Pricing  |  {customer}  |  {datetime.today().strftime('%d %b %Y')}"
    t.font      = _font(bold=True, size=13, color="FFFFFF")
    t.fill      = _fill("1F4E79")
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{col_letter}2")
    s = ws["A2"]
    s.value     = f"Base Monthly: ${pricing['total_monthly_usd']:,.2f}  |  Annual Y1: ${pricing['total_annual_usd']:,.2f}  |  3-Year: ${pricing['total_3year_usd']:,.2f}"
    s.font      = _font(italic=True, size=10, color="FFFFFF")
    s.fill      = _fill("2E75B6")
    s.alignment = _center()
    ws.row_dimensions[2].height = 20

    year_labels = [f"Year {y+1}" for y in range(years)]
    headers = ["Category", "Service / Role", "Instance Type", "Nodes", "Base Monthly (All Nodes)", "Annual Y1"] + year_labels[1:] + ["5-Year Total"]
    _header_row(ws, 3, headers)
    ws.row_dimensions[3].height = 32

    num_cols = set(range(5, 5 + years + 1 + 1))
    row = 4
    prev_cat = None

    priced_roles = sorted(pricing.get("priced_roles", []), key=lambda r: r.get("category", ""))

    for i, r in enumerate(priced_roles):
        cat     = r.get("category", "Other")
        base    = r.get("monthly_usd", 0)
        if base == 0:
            continue

        if cat != prev_cat:
            ws.merge_cells(f"A{row}:{col_letter}{row}")
            c = ws.cell(row=row, column=1, value=f"  {cat}")
            c.font      = _font(bold=True, size=10, color="FFFFFF")
            c.fill      = _fill("2E75B6")
            c.alignment = _left()
            c.border    = _border()
            ws.row_dimensions[row].height = 18
            row += 1
            prev_cat = cat

        year_costs = [round(base * 12 * ((1 + INFLATION) ** y), 2) for y in range(years)]
        five_yr    = round(sum(year_costs), 2)

        vals = [cat, r.get("label", "—"), r.get("instance_type", "—"), r.get("nodes", "—"), base] + year_costs + [five_yr]
        _data_row(ws, row, vals, alt=(i % 2 == 0), num_cols=num_cols)
        ws.row_dimensions[row].height = 18
        row += 1

    # Total row
    base_monthly = pricing["total_monthly_usd"]
    year_costs   = [round(base_monthly * 12 * ((1 + INFLATION) ** y), 2) for y in range(years)]
    five_yr      = round(sum(year_costs), 2)
    _data_row(ws, row, ["TOTAL", "", "", "", base_monthly] + year_costs + [five_yr], total=True, num_cols=num_cols)

    _set_col_widths(ws, [26, 40, 18, 10, 22] + [16] * years + [16])




# ── Sheet: ClickHouse Cluster Sizing ───────────────────────────────────────

def _build_clickhouse_sheet(wb: Workbook, ch_sizing: dict, customer: str):
    """
    Two-table ClickHouse sheet:
      Table 1: DB cluster  (Shard, Node, vCPU, RAM, Total RAM/shard, Total Cores/shard, Storage)
      Table 2: Keeper cluster (Node, vCPU, RAM, Total RAM, Total Cores, Storage)
    """
    ws = wb.create_sheet("ClickHouse Sizing")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    db_cl  = ch_sizing.get("db_cluster",  {})
    kp_cl  = ch_sizing.get("keeper_cluster", {})
    summ   = ch_sizing.get("summary", {})

    # ── Title block ──────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value     = f"ClickHouse OLAP Cluster Sizing  |  {customer}  |  {datetime.today().strftime('%d %b %Y')}"
    t.font      = _font(bold=True, size=13, color="FFFFFF")
    t.fill      = _fill("1F4E79")
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = (
        f"Total Nodes: {summ.get('total_nodes', 0)}"
        f"   |   Shards: {summ.get('num_shards', 0)}"
        f"   |   Total vCPU: {summ.get('total_vcpu', 0)}"
        f"   |   Total RAM: {summ.get('total_ram_gb', 0)} GB"
        f"   |   Total Storage: {summ.get('total_storage_gb', 0):,} GB"
        f"   |   Analytics Data: {ch_sizing.get('ch_data_gb', 0):,.0f} GB"
    )
    s.font      = _font(italic=True, size=10, color="FFFFFF")
    s.fill      = _fill("2E75B6")
    s.alignment = _center()
    ws.row_dimensions[2].height = 18

    row = 3

    # ── Section 1: ClickHouse DB Cluster ──────────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1,
                value="  ClickHouse DB Cluster — Sharded + Replicated (Self-Hosted EC2)")
    c.font      = _font(bold=True, size=10, color="FFFFFF")
    c.fill      = _fill("1F4E79")
    c.alignment = _left()
    c.border    = _border()
    ws.row_dimensions[row].height = 20
    row += 1

    db_headers = [
        "Shard", "Node (Replica)",
        "CPU (vCPU)", "Memory (GB)",
        "Total Memory / Shard (GB)", "Total Cores / Shard",
        "SSD Storage / Node (GB)", "Role",
    ]
    _header_row(ws, row, db_headers)
    ws.row_dimensions[row].height = 28
    row += 1

    n_shards = db_cl.get("num_shards", 1)
    shard_ram   = db_cl.get("total_ram_gb", 0) // n_shards if n_shards else 0
    shard_vcpu  = db_cl.get("total_vcpu",  0) // n_shards if n_shards else 0

    for i, node in enumerate(db_cl.get("nodes", [])):
        vals = [
            f"Shard {node['shard']}",
            f"Replica {node['replica']}",
            node["vcpu_per_node"],
            node["ram_per_node"],
            shard_ram,
            shard_vcpu,
            node["storage_per_node_gb"],
            "ClickHouse DB (OLAP)",
        ]
        _data_row(ws, row, vals, alt=(i % 2 == 0))
        ws.cell(row=row, column=7).number_format = "#,##0"
        ws.row_dimensions[row].height = 18
        row += 1

    # DB cluster total
    _data_row(ws, row, [
        "TOTAL", f"{db_cl.get('total_nodes', 0)} nodes",
        db_cl.get("total_vcpu", 0), db_cl.get("total_ram_gb", 0),
        "", "",
        db_cl.get("total_nodes", 0) * db_cl.get("storage_per_node_gb", 0),
        "",
    ], total=True)
    ws.cell(row=row, column=7).number_format = "#,##0"
    row += 2

    # ── Section 2: Keeper Cluster ─────────────────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1,
                value="  ClickHouse Keeper Cluster — 3-Node Raft Quorum (Coordination)")
    c.font      = _font(bold=True, size=10, color="FFFFFF")
    c.fill      = _fill("1F4E79")
    c.alignment = _left()
    c.border    = _border()
    ws.row_dimensions[row].height = 20
    row += 1

    kp_headers = [
        "Node Type", "Node Label",
        "CPU (vCPU)", "Memory (GB)",
        "Total Memory / Cluster (GB)", "Total Cores / Cluster",
        "SSD Storage / Node (GB)", "Role",
    ]
    _header_row(ws, row, kp_headers)
    ws.row_dimensions[row].height = 28
    row += 1

    for i, node in enumerate(kp_cl.get("nodes", [])):
        lbl = node.get("label", f"Keeper {i+1}")
        short_lbl = lbl.split("—")[-1].strip() if "—" in lbl else lbl
        vals = [
            "Keeper Node",
            short_lbl,
            node["vcpu_per_node"],
            node["ram_per_node"],
            kp_cl.get("total_ram_gb", 0),
            kp_cl.get("total_vcpu", 0),
            node["storage_per_node_gb"],
            "ClickHouse Keeper (Raft Quorum)",
        ]
        _data_row(ws, row, vals, alt=(i % 2 == 0))
        ws.row_dimensions[row].height = 18
        row += 1

    # Keeper total
    _data_row(ws, row, [
        "TOTAL", f"{kp_cl.get('total_nodes', 0)} nodes",
        kp_cl.get("total_vcpu", 0), kp_cl.get("total_ram_gb", 0),
        "", "",
        kp_cl.get("total_nodes", 0) * kp_cl.get("storage_per_node_gb", 0),
        "",
    ], total=True)
    row += 2

    # ── Sizing assumptions note ───────────────────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1, value="  Sizing Notes")
    c.font = _font(bold=True, size=10, color="FFFFFF")
    c.fill = _fill("1F4E79")
    c.alignment = _left()
    c.border = _border()
    row += 1

    notes = [
        ("Architecture",         "Self-hosted on EC2 for both SaaS and On-Prem (no managed ClickHouse)"),
        ("Analytics Data Volume", f"{ch_sizing.get('ch_data_gb', 0):,.0f} GB  ({ch_sizing.get('ch_data_multiplier', 2.0)}× transactional DB × volume factor {ch_sizing.get('volume_factor', 1.0)}×)"),
        ("DB Node Type",         "Memory-Intensive (r5/r6a family) — optimised for columnar merge I/O"),
        ("Keeper Node Type",     "General Purpose (m5/c6i family) — Raft quorum metadata coordination"),
        ("Storage Type",         "EBS gp3 for DB nodes and Keeper"),
        ("HA / Replication",     f"{db_cl.get('num_shards', 1)} shard(s) × {db_cl.get('replicas_per_shard', 2)} replicas — intra-cluster replication; Keeper ensures quorum"),
    ]
    for i, (k, v) in enumerate(notes):
        for col, val in enumerate([k, v, "", "", "", "", "", ""], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = _font(size=10, color="0000FF" if col == 2 else "000000")
            c.fill      = _fill("EBF3FB") if i % 2 == 0 else _fill("FFFFFF")
            c.alignment = _left()
            c.border    = _border()
        row += 1

    _set_col_widths(ws, [20, 28, 12, 14, 26, 22, 24, 30])


# ── Sheet: Pre-Prod / SIT / UAT ──────────────────────────────────────────

# ── Sheet: Environment Pricing (Pre-Prod / DR) ───────────────────────────

def _build_env_pricing_sheet(
    wb: Workbook,
    env_data: dict,
    sheet_name: str,
    title: str,
    customer: str,
    include_forecast: bool = False,
    row_multiplier: int = 1,
):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    years = 5
    INFLATION = 0.04

    # Title
    total_cols = (5 + years + 1) if include_forecast else 8
    col_letter = get_column_letter(total_cols)
    ws.merge_cells(f"A1:{col_letter}1")
    t = ws["A1"]
    t.value     = f"{title} Pricing  |  {customer}  |  {datetime.today().strftime('%d %b %Y')}"
    t.font      = _font(bold=True, size=13, color="FFFFFF")
    t.fill      = _fill("1F4E79")
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells(f"A2:{col_letter}2")
    s = ws["A2"]
    s.value     = f"Monthly: ${env_data.get('monthly_usd',0):,.2f}   |   Annual: ${env_data.get('annual_usd',0):,.2f}"
    if include_forecast:
        five_yr = env_data.get("five_year_forecast", {}).get("five_year_total", 0)
        s.value += f"   |   5-Year Total: ${five_yr:,.2f}"
    s.font      = _font(italic=True, size=10, color="FFFFFF")
    s.fill      = _fill("2E75B6")
    s.alignment = _center()
    ws.row_dimensions[2].height = 18

    # Headers
    headers = ["Category", "Service / Role", "Instance Type", "Nodes", "Base Monthly (All Nodes)"]
    if include_forecast:
        headers += [f"Annual Y1"] + [f"Year {y+1}" for y in range(1, years)] + ["5-Year Total"]
    else:
        headers += ["Annual USD", "vCPU / Node", "RAM / Node (GB)"]

    _header_row(ws, 3, headers)
    ws.row_dimensions[3].height = 28

    num_cols = set(range(5, total_cols + 1))
    row = 4
    prev_cat = None

    priced_roles = sorted(env_data.get("priced_roles", []), key=lambda r: r.get("category", ""))

    for i, r in enumerate(priced_roles):
        cat  = r.get("category", "Other")
        # Apply multiplier: ×2 for SIT+UAT (2 environments), ×1 for DR
        base = round((r.get("monthly_usd", 0) or 0) * row_multiplier, 2)
        if base == 0:
            continue

        if cat != prev_cat:
            ws.merge_cells(f"A{row}:{col_letter}{row}")
            c = ws.cell(row=row, column=1, value=f"  {cat}")
            c.font = _font(bold=True, size=10, color="FFFFFF")
            c.fill = _fill("2E75B6"); c.alignment = _left(); c.border = _border()
            row += 1
            prev_cat = cat

        if include_forecast:
            year_costs = [round(base * 12 * ((1 + INFLATION) ** y), 2) for y in range(years)]
            five_yr    = round(sum(year_costs), 2)
            vals = [cat, r.get("label", "—"), r.get("instance_type", "—"), r.get("nodes", "—"), base] + year_costs + [five_yr]
        else:
            vals = [cat, r.get("label", "—"), r.get("instance_type", "—"), r.get("nodes", "—"), base,
                    base * 12, r.get("vcpu_per_node", "—"), r.get("ram_per_node", "—")]

        _data_row(ws, row, vals, alt=(i % 2 == 0), num_cols=num_cols)
        ws.row_dimensions[row].height = 18
        row += 1

    # Total row
    base_monthly = env_data.get("monthly_usd", 0)
    if include_forecast:
        year_costs   = [round(base_monthly * 12 * ((1 + INFLATION) ** y), 2) for y in range(years)]
        five_yr      = round(sum(year_costs), 2)
        vals = ["TOTAL", "", "", "", base_monthly] + year_costs + [five_yr]
    else:
        vals = ["TOTAL", "", "", "", base_monthly, base_monthly * 12, "", ""]

    _data_row(ws, row, vals, total=True, num_cols=num_cols)

    if include_forecast:
        _set_col_widths(ws, [22, 40, 22, 10, 22] + [16] * years + [16])
    else:
        _set_col_widths(ws, [22, 45, 22, 10, 22, 18, 14, 16])


# ── Sheet: PUPM Summary (Year-by-Year) ───────────────────────────────────

def _build_pupm_sheet(wb: Workbook, pricing: dict, env_pricing: dict | None,
                      metrics: dict, customer: str, ai_sizing: dict = None,
                      cloud: str = "AWS"):
    """
    Builds the PUPM Summary sheet replicating the reference workbook format:

    Rows (per year column):
      BusinessNext section:
        - Production DC (monthly)
        - Production DR (monthly, if applicable)
        - Pre-Prod       (monthly, if applicable)
        - UAT            (blank — not separately priced)
        - SIT            (blank — not separately priced)
        - Performance Testing  [One-Time: $5,000]
        - Migration/Data Bootup [One-Time: $5,000]  ← Optional Component
      Security section:
        - Airtel SOC             $500/mo
        - SOC Machines           $416.67/mo  (Rs 30,000/unit/yr)
        - Req 4 – Antivirus & EDR  $36.11/mo  (Rs 2,600/unit/yr)
        - Req 6 – Data Discovery   $55.56/mo  (Rs 4,000/unit/yr)
        - Req 7 – DLP              $0
      ── Calculations ──
        Total Usage     = sum of all monthly costs
        Business Support = 3.91% of Total Usage
        Total Platform cost (Yearly) = (Total Usage + Business Support) × 12
        Buffer          = 5% of Total Platform cost
        Total cost (AWS) = Total Platform cost + Buffer
        Managed Services = 30% of Total cost AWS
        Total cost      = Total cost AWS + Managed Services
        Discount        = 11%
        Discounted cost = Total cost × (1 − 0.11)
        ── PUPM ──
        Users           = named users for that year
        PUPM            = Discounted cost / 12 / Users
    """
    ws = wb.create_sheet("PUPM Summary")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "E4"

    INFLATION  = pricing.get("inflation_rate", 0.04)
    num_years  = 5

    # ── Derive yearly Production DC costs with inflation ──────────────────
    base_monthly  = pricing.get("total_monthly_usd", 0)

    # DR / Pre-Prod monthly
    dr_monthly     = 0.0
    preprod_monthly = 0.0
    if env_pricing:
        dr_data = env_pricing.get("dr")
        pp_data = env_pricing.get("preprod_sit_uat")
        if dr_data:
            dr_monthly = dr_data.get("monthly_usd", 0.0)
        if pp_data:
            preprod_monthly = pp_data.get("monthly_usd", 0.0)

    # Named users per year (grow at 5% YOY per app.py YOY settings)
    named_users_y1 = metrics.get("total_named_users",
                     metrics.get("named_users",
                     int(metrics.get("mobile_users", 0) / 0.30 * 1.0))) or 8500

    # ── One-time costs from user inputs (defaulting to $5000/$5000/$1000) ──
    ot_perf = float(metrics.get("one_time_perf_testing", 5000) or 5000)
    ot_migr = float(metrics.get("one_time_migration",    5000) or 5000)
    ot_ms   = float(metrics.get("one_time_managed_svc",  1000) or 1000)

    # Build per-year data
    years = []
    for y in range(1, num_years + 1):
        mult     = (1 + INFLATION) ** (y - 1)
        prod_dc  = round(base_monthly * mult, 2)
        prod_dr  = round(dr_monthly * mult, 2)  if dr_monthly  else 0.0
        preprod  = round(preprod_monthly * mult, 2) if preprod_monthly else 0.0
        users_y  = int(named_users_y1 * (1.05 ** (y - 1)))

        # Fixed security costs (INR-based, fixed in USD)
        airtel_soc  = 500.00
        soc_mach    = round(5000  / 12, 4)   # $5000/yr → $416.67/mo
        req4        = round(433.33 / 12, 4)  # Rs 2,600/unit/yr
        req6        = round(666.67 / 12, 4)  # Rs 4,000/unit/yr
        req7        = 0.0

        # One-time costs: only in Year 1, zero for subsequent years
        one_time_perf  = ot_perf if y == 1 else 0.0
        one_time_migr  = ot_migr if y == 1 else 0.0
        one_time_ms    = ot_ms   if y == 1 else 0.0
        one_time_total = one_time_perf + one_time_migr + one_time_ms

        # AI Services monthly cost (compute + storage + bedrock), inflated
        ai_base_monthly = 0.0
        if ai_sizing and ai_sizing.get("enabled"):
            cs = ai_sizing.get("combined_summary", {})
            bedrock_mo = cs.get("bedrock_monthly", 3000)
            # Approximate EC2 compute: $0.768/hr × 730 hrs × total_nodes (conservative avg)
            HOURS_PER_MONTH = 730
            NODE_HOURLY_AVG = 0.576   # blended avg across all AI types
            nodes = cs.get("total_worker_nodes", 0)
            stor_gb = cs.get("total_storage_gb", 0)
            compute = nodes * NODE_HOURLY_AVG * HOURS_PER_MONTH
            storage = stor_gb * 0.08   # EBS GP3
            ai_base_monthly = round(compute + storage + bedrock_mo, 2)

        total_usage = (prod_dc + prod_dr + preprod
                       + airtel_soc + soc_mach + req4 + req6 + req7
                       + round(ai_base_monthly * mult, 2))

        BUS_SUPPORT_PCT  = 0.039127
        BUFFER_PCT       = 0.05
        MANAGED_SVC_PCT  = 0.30
        DISCOUNT         = 0.11

        business_support   = round(total_usage * BUS_SUPPORT_PCT, 4)
        total_platform_yr  = round((total_usage + business_support) * 12, 4)
        buffer             = round(total_platform_yr * BUFFER_PCT, 4)
        total_aws          = round(total_platform_yr + buffer, 4)
        managed_svc        = round(total_aws * MANAGED_SVC_PCT, 4)
        total_cost         = round(total_aws + managed_svc + one_time_total, 4)
        discounted_cost    = round(total_cost * (1 - DISCOUNT), 4)
        pupm               = round(discounted_cost / 12 / users_y, 4) if users_y else 0

        years.append({
            "year":             y,
            "prod_dc":          prod_dc,
            "prod_dr":          prod_dr,
            "preprod":          preprod,
            "ai_monthly":       round(ai_base_monthly * mult, 2),
            "airtel_soc":       airtel_soc,
            "soc_mach":         soc_mach,
            "req4":             req4,
            "req6":             req6,
            "req7":             req7,
            "one_time_perf":    one_time_perf,
            "one_time_migr":    one_time_migr,
            "total_usage":      total_usage,
            "business_support": business_support,
            "total_platform_yr":total_platform_yr,
            "buffer":           buffer,
            "total_aws":        total_aws,
            "managed_svc":      managed_svc,
            "managed_svc_ot":   one_time_ms,
            "total_cost":       total_cost,
            "one_time_total":   one_time_total,
            "discounted_cost":  discounted_cost,
            "users":            users_y,
            "pupm":             pupm,
        })

    # ── Column layout ─────────────────────────────────────────────────────
    # A=Category, B=Line Item, C=Comments, D=One-Time ($), E..I = Y1..Y5 ($/month)
    _set_col_widths(ws, [16, 48, 30, 16, 16, 16, 16, 16, 16])

    # ── Title ─────────────────────────────────────────────────────────────
    last_col = get_column_letter(4 + num_years)
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value     = f"PUPM Summary – {num_years}-Year Cost Analysis  |  {customer}  |  {datetime.today().strftime('%d %b %Y')}  |  Inflation: {INFLATION*100:.1f}%/yr"
    t.font      = _font(bold=True, size=13, color="FFFFFF")
    t.fill      = _fill("1F4E79")
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    # ── Sub-title with PUPM range ─────────────────────────────────────────
    ws.merge_cells(f"A2:{last_col}2")
    pupm_range = f"PUPM Y1: ${years[0]['pupm']:,.2f}  →  PUPM Y{num_years}: ${years[-1]['pupm']:,.2f}   |   Discount: 11%  |  Managed Services: 30%"
    s = ws["A2"]
    s.value     = pupm_range
    s.font      = _font(italic=True, size=10, color="FFFFFF")
    s.fill      = _fill("2E75B6")
    s.alignment = _center()
    ws.row_dimensions[2].height = 18

    # ── Column headers ────────────────────────────────────────────────────
    hdrs = ["Category", "Line Item", "Comments", "One-Time Cost ($)"]
    hdrs += [f"Year {y}\n($/month)" for y in range(1, num_years + 1)]
    _header_row(ws, 3, hdrs)
    ws.row_dimensions[3].height = 36

    # ── Helper to write a section header row ─────────────────────────────
    def section_hdr(row_num, title, fill_hex="2E75B6"):
        ws.merge_cells(f"A{row_num}:{last_col}{row_num}")
        c = ws.cell(row=row_num, column=1, value=f"  {title}")
        c.font      = _font(bold=True, size=10, color="FFFFFF")
        c.fill      = _fill(fill_hex)
        c.alignment = _left()
        c.border    = _border()
        ws.row_dimensions[row_num].height = 20

    # ── Helper to write a data row ────────────────────────────────────────
    MONEY = '$#,##0.00'
    PCTFMT = '0.00%'

    def data_row(row_num, category, label, comment, one_time,
                 yr_vals, alt=False, bold=False, total=False,
                 is_pct=False, is_pupm=False):
        fill = TOTAL_FILL if total else (_fill("EBF3FB") if alt else _fill("FFFFFF"))
        if is_pupm:
            fill = _fill("D6E4F0")  # distinct light-blue highlight for PUPM

        cells_vals = [category, label, comment, one_time] + list(yr_vals)
        for col_i, val in enumerate(cells_vals, 1):
            c = ws.cell(row=row_num, column=col_i, value=val)
            c.font      = _font(bold=bold or total or is_pupm, size=10,
                                color="000000" if not (bold and col_i == 1) else "000000")
            c.fill      = fill
            c.border    = _border()
            c.alignment = _left() if col_i <= 3 else _center()
            if col_i == 4 and isinstance(val, (int, float)) and val:
                c.number_format = MONEY
            if col_i >= 5:
                if is_pct:
                    c.number_format = PCTFMT
                elif is_pupm:
                    c.number_format = '$#,##0.0000'
                elif isinstance(val, (int, float)):
                    c.number_format = MONEY

    def blank_row(row_num):
        for col_i in range(1, 5 + num_years):
            c = ws.cell(row=row_num, column=col_i, value="")
            c.fill   = _fill("FFFFFF")
            c.border = _border()
        ws.row_dimensions[row_num].height = 6

    # ── Build rows ────────────────────────────────────────────────────────
    r = 4

    # ── BUSINESSNEXT section ──────────────────────────────────────────────
    section_hdr(r, "BusinessNext"); r += 1

    data_row(r, "BusinessNext", "Production DC", "",
             "", [y["prod_dc"] for y in years], alt=False); r += 1

    if any(y["prod_dr"] for y in years):
        data_row(r, "", "Production DR", "",
                 "", [y["prod_dr"] for y in years], alt=True); r += 1

    if any(y["preprod"] for y in years):
        data_row(r, "", "Pre-Prod", "",
                 "", [y["preprod"] for y in years], alt=False); r += 1

    data_row(r, "", "UAT", "",  "", [""] * num_years, alt=True);  r += 1
    data_row(r, "", "SIT", "",  "", [""] * num_years, alt=False); r += 1

    data_row(r, "", "Performance Testing",
             "One-time — Year 1 only", years[0]["one_time_perf"],
             [""] * num_years, alt=True); r += 1

    data_row(r, "", "Migration / Data Bootup",
             "One-time — Year 1 only", years[0]["one_time_migr"],
             [""] * num_years, alt=False); r += 1

    # ── AI Services rows (only when AI is enabled) ────────────────────────
    if ai_sizing and ai_sizing.get("enabled"):
        blank_row(r); r += 1
        section_hdr(r, "AI Services", fill_hex="0F3460"); r += 1

        envs = ai_sizing.get("environments", {})
        _type_labels = {"predictive": "Predictive AI", "genai": "GenAI", "agentic": "Agentic AI"}
        for ai_type, envs_data in envs.items():
            lbl = _type_labels.get(ai_type, ai_type.upper())
            prod_env = envs_data.get("prod", {})
            if not prod_env:
                continue
            cluster = prod_env.get("cluster", {})
            nodes   = cluster.get("worker_nodes", 0)
            cores   = cluster.get("total_cores", 0)
            ram     = cluster.get("total_ram_gb", 0)
            comment = f"{nodes} workers · {cores} vCPU · {ram} GB RAM"
            # Compute this AI type's share of AI monthly cost (proportional to nodes)
            HOURS_PER_MONTH = 730
            NODE_HOURLY_AVG = 0.576
            stor_gb = cluster.get("total_storage_gb", 0)
            type_monthly = round(
                nodes * NODE_HOURLY_AVG * HOURS_PER_MONTH
                + stor_gb * 0.08, 2
            )
            data_row(r, "AI Services", f"{lbl} (Production)",
                     comment, "",
                     [round(type_monthly * ((1 + years[0].get('pupm', 0.04) / years[0]['pupm'] if years[0]['pupm'] else 1) ** max(0, yi - 1)), 2)
                      if False else round(type_monthly * ((1.04) ** yi), 2)
                      for yi in range(num_years)],
                     alt=(list(envs.keys()).index(ai_type) % 2 == 1)); r += 1

        # Bedrock API cost row
        cs = ai_sizing.get("combined_summary", {})
        bedrock_mo = cs.get("bedrock_monthly", 3000)
        data_row(r, "", "AWS Bedrock API (Managed LLM)",
                 "Token-based — no GPU hardware", "",
                 [round(bedrock_mo * ((1.04) ** yi), 2) for yi in range(num_years)],
                 alt=True); r += 1

        # AI Total row
        data_row(r, "", "AI Services Total (Monthly)",
                 "Included in Total Usage below", "",
                 [y["ai_monthly"] for y in years],
                 bold=True, total=True); r += 1

    # ── SECURITY section ──────────────────────────────────────────────────
    section_hdr(r, "Security"); r += 1

    data_row(r, "Security", "Airtel SOC", "Infra for SOC machines",
             "", [y["airtel_soc"] for y in years], alt=False); r += 1
    data_row(r, "", "SOC Machines", "Rs. 30,000/unit/year",
             "", [y["soc_mach"] for y in years], alt=True); r += 1
    data_row(r, "", "Requirement 4 – Antivirus & EDR", "Rs. 2,600/unit/year",
             "", [y["req4"] for y in years], alt=False); r += 1
    data_row(r, "", "Requirement 6 – Data Discovery & Classification", "Rs. 4,000/unit/year",
             "", [y["req6"] for y in years], alt=True); r += 1
    data_row(r, "", "Requirement 7 – DLP", "",
             "", [y["req7"] for y in years], alt=False); r += 1

    blank_row(r); r += 1

    # ── CALCULATIONS section ──────────────────────────────────────────────
    section_hdr(r, "Calculations", fill_hex="1F4E79"); r += 1

    data_row(r, "", "Total Usage", "",
             "", [y["total_usage"] for y in years],
             bold=True, total=True); r += 1

    data_row(r, "", "Business Support (3.91% of Total Usage)", "",
             "", [y["business_support"] for y in years], alt=True); r += 1

    data_row(r, "", "Total Platform cost (Yearly cost)", "",
             sum(y["one_time_perf"] + y["one_time_migr"] for y in years[:1]),
             [y["total_platform_yr"] for y in years],
             bold=True); r += 1

    data_row(r, "", "Buffer (5%)", "",
             "", [y["buffer"] for y in years], alt=True); r += 1

    data_row(r, "", f"Total cost ({cloud})", "",
             sum(y["one_time_perf"] + y["one_time_migr"] for y in years[:1]),
             [y["total_aws"] for y in years],
             bold=True, total=True); r += 1

    data_row(r, "", f"Managed Services (30% of {cloud} cost)", "",
             years[0]["managed_svc_ot"],
             [y["managed_svc"] for y in years], alt=True); r += 1

    data_row(r, "", "Total cost", "",
             sum(y["one_time_total"] for y in years[:1]),
             [y["total_cost"] for y in years],
             bold=True, total=True); r += 1

    data_row(r, "", "Discount", "",
             "", [0.11] * num_years, is_pct=True, alt=True); r += 1

    data_row(r, "", "Discounted cost", "",
             sum(y["one_time_total"] for y in years[:1]),
             [y["discounted_cost"] for y in years],
             bold=True, total=True); r += 1

    blank_row(r); r += 1

    # ── PUPM section ──────────────────────────────────────────────────────
    section_hdr(r, "PUPM Calculation", fill_hex="1F4E79"); r += 1

    data_row(r, "", "Named Users (YOY +5%)", "",
             "", [y["users"] for y in years], alt=False); r += 1

    # The star row
    data_row(r, "", "PUPM  (Price per User per Month)", "",
             "", [y["pupm"] for y in years],
             is_pupm=True, bold=True); r += 1

    ws.row_dimensions[r - 1].height = 26

    blank_row(r); r += 1

    # ── 5-Year PUPM trend table (visual summary) ──────────────────────────
    section_hdr(r, "5-Year PUPM Trend", fill_hex="2E75B6"); r += 1

    trend_hdrs = ["", "Metric"] + [f"Year {y}" for y in range(1, num_years + 1)] + ["5-Yr Avg"]
    for col_i, val in enumerate(trend_hdrs, 1):
        c = ws.cell(row=r, column=col_i, value=val)
        c.font = _font(bold=True, size=10, color="FFFFFF")
        c.fill = _fill("2E75B6"); c.border = _border(); c.alignment = _center()
    r += 1

    trend_rows = [
        ("Monthly Cost (Prod DC)",  [y["prod_dc"]  for y in years]),
        ("Total Monthly Usage",     [y["total_usage"] for y in years]),
        ("Discounted Cost (Annual)",[y["discounted_cost"] for y in years]),
        ("Named Users",             [y["users"] for y in years]),
        ("PUPM",                    [y["pupm"]  for y in years]),
    ]
    for t_i, (label, vals) in enumerate(trend_rows):
        avg = sum(vals) / len(vals)
        row_vals = ["", label] + vals + [avg]
        is_p = (label == "PUPM")
        fill = _fill("D6E4F0") if is_p else (_fill("EBF3FB") if t_i % 2 == 0 else _fill("FFFFFF"))
        for col_i, val in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=col_i, value=val)
            c.font   = _font(bold=is_p, size=10)
            c.fill   = fill; c.border = _border()
            c.alignment = _left() if col_i == 2 else _center()
            if col_i >= 3 and isinstance(val, (int, float)):
                c.number_format = '$#,##0.0000' if is_p else MONEY
        r += 1

    # ── Footnotes ─────────────────────────────────────────────────────────
    r += 1
    notes = [
        "Notes:",
        "• PUPM = Price Per User Per Month = Discounted Annual Cost ÷ 12 ÷ Named Users",
        "• Business Support: 3.91% of Total Monthly Usage",
        "• Buffer: 5% of Total Platform Annual Cost",
        f"• Managed Services: 30% of Total {cloud} Cost (Annual)",
        "• Discount: 11% applied to Total Cost",
        "• Production DC cost grows at inflation rate year-on-year",
        "• Named Users grow at 5% YOY",
        "• Performance Testing & Migration are one-time costs in Year 1 only",
    ]
    for note in notes:
        ws.merge_cells(f"A{r}:{last_col}{r}")
        c = ws.cell(row=r, column=1, value=note)
        c.font      = _font(italic=(note != "Notes:"), bold=(note == "Notes:"), size=9, color="595959")
        c.fill      = _fill("F2F2F2")
        c.alignment = _left()
        c.border    = _border()
        ws.row_dimensions[r].height = 16
        r += 1


# ── AI Services Sheets ────────────────────────────────────────────────────

# Colour palette for AI sheets
_AI_HDR_FILL  = _fill("1A1A2E")   # deep navy
_AI_SEC_FILL  = _fill("16213E")   # mid navy
_AI_TYPE_FILLS = {
    "agentic":   _fill("0F3460"),   # dark blue
    "predictive":_fill("533483"),   # purple
    "genai":     _fill("E94560"),   # red
}
_AI_ALT_FILL  = _fill("EEF2FF")   # very light lavender
_AI_WHT_FILL  = _fill("FFFFFF")

_AI_TYPE_LABELS = {
    "predictive": "📈 Predictive AI",
    "genai":      "💬 GenAI",
    "agentic":    "🤖 Agentic AI",
}


def _build_ai_sizing_sheet(wb: Workbook, ai_sizing: dict, customer: str):
    """
    Creates one separate worksheet per enabled AI service type
    (Predictive AI, Generative AI, Agentic AI), each showing:
      Table 1: Service pod requirements
      Table 2: K8s cluster configuration
      Table 3: Database requirements
    """
    HDR_FILL = _AI_HDR_FILL
    SEC_FILL = _AI_SEC_FILL
    environments = ai_sizing.get("environments", {})

    for ai_type, envs_data in environments.items():
        type_label = _AI_TYPE_LABELS.get(ai_type, ai_type.upper())
        type_fill  = _AI_TYPE_FILLS.get(ai_type, SEC_FILL)
        sheet_name = f"{type_label} Sizing"[:31]  # Excel sheet name max 31 chars

        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        # Title row
        ws.merge_cells("A1:K1")
        t = ws["A1"]
        t.value     = f"{type_label} Infrastructure Sizing — {customer}"
        t.font      = _font(bold=True, size=13, color="FFFFFF")
        t.fill      = type_fill
        t.alignment = _center()
        ws.row_dimensions[1].height = 28

        # Bedrock note for GenAI
        bedrock = ai_sizing.get("bedrock_monthly", 3000)
        ws.merge_cells("A2:K2")
        n = ws["A2"]
        if ai_type in ("genai", "generative"):
            n.value = (
                f"AWS Bedrock (Managed LLM Inference) — No self-hosted GPU required  |  "
                f"Bedrock estimated monthly cost: ${bedrock:,.0f}  |  "
                f"Source: Maybank Hardware Sizing AI Template"
            )
        else:
            n.value = f"{type_label} — GPU compute sizing  |  Source: Maybank Hardware Sizing AI Template"
        n.font      = _font(size=9, italic=True, color="FFFFFF")
        n.fill      = SEC_FILL
        n.alignment = _left()
        ws.row_dimensions[2].height = 16

        r = 3

        for env_key, env_data in envs_data.items():
            env_label = env_data.get("env_label", env_key)

            # ── Type + Env header ──────────────────────────────────────────
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
            h = ws.cell(row=r, column=1,
                        value=f"  {type_label}  —  {env_label}")
            h.font      = _font(bold=True, size=11, color="FFFFFF")
            h.fill      = type_fill
            h.alignment = _left()
            h.border    = _border()
            ws.row_dimensions[r].height = 22
            r += 1

            # ── Table 1: Services ──────────────────────────────────────────
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
            sh = ws.cell(row=r, column=1, value="  DataNext Services — Pod Requirements")
            sh.font      = _font(bold=True, size=10, color="FFFFFF")
            sh.fill      = _fill("2E4057")
            sh.alignment = _left()
            sh.border    = _border()
            ws.row_dimensions[r].height = 18
            r += 1

            svc_hdrs = ["#", "Service Name", "Category",
                        "CPU/Pod", "RAM/Pod (GB)", "Pods",
                        "Total CPU", "Total RAM (GB)", "Storage (GB)",
                        "Persistent", ""]
            for ci, h in enumerate(svc_hdrs, 1):
                c = ws.cell(row=r, column=ci, value=h)
                c.font      = _font(bold=True, size=9, color="FFFFFF")
                c.fill      = _fill("3D5A80")
                c.alignment = _center()
                c.border    = _border()
            ws.row_dimensions[r].height = 16
            r += 1

            services = env_data.get("services", [])
            for i, svc in enumerate(services):
                fill = _AI_ALT_FILL if i % 2 else _AI_WHT_FILL
                total_cpu = svc["cpu_per_pod"] * svc["pods"]
                total_ram = svc["ram_per_pod"] * svc["pods"]
                stor = svc.get("storage_gb", 0)
                row_vals = [i+1, svc["name"], svc.get("category",""),
                            svc["cpu_per_pod"], svc["ram_per_pod"], svc["pods"],
                            total_cpu, total_ram,
                            stor if stor else "—",
                            "Yes" if stor else "No", ""]
                for ci, v in enumerate(row_vals, 1):
                    c = ws.cell(row=r, column=ci, value=v)
                    c.font      = _font(size=9)
                    c.fill      = fill
                    c.alignment = _center() if ci > 1 else _left()
                    c.border    = _border()
                ws.row_dimensions[r].height = 15
                r += 1

            # Totals row for services
            svc_total_cpu = sum(s["cpu_per_pod"]*s["pods"] for s in services)
            svc_total_ram = sum(s["ram_per_pod"]*s["pods"] for s in services)
            svc_total_stor = sum(s.get("storage_gb",0) for s in services)
            t_vals = ["", "TOTAL", "", "", "", sum(s["pods"] for s in services),
                      svc_total_cpu, svc_total_ram,
                      svc_total_stor if svc_total_stor else "—", "", ""]
            for ci, v in enumerate(t_vals, 1):
                c = ws.cell(row=r, column=ci, value=v)
                c.font      = _font(bold=True, size=9)
                c.fill      = TOTAL_FILL
                c.alignment = _center()
                c.border    = _border()
            ws.row_dimensions[r].height = 16
            r += 2

            # ── Table 2: Kubernetes Cluster Configuration ──────────────────
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
            sh2 = ws.cell(row=r, column=1, value="  Kubernetes Cluster Configuration")
            sh2.font      = _font(bold=True, size=10, color="FFFFFF")
            sh2.fill      = _fill("2E4057")
            sh2.alignment = _left()
            sh2.border    = _border()
            ws.row_dimensions[r].height = 18
            r += 1

            cluster = env_data.get("cluster", {})
            kube_hdrs = ["Node Type", "Number", "Phys. Cores/Node",
                         "RAM/Node (GB)", "Total RAM (GB)", "Total Cores",
                         "SSD Storage/Node (GB)", "Total Storage (GB)",
                         "Processor", "GPU", ""]
            for ci, h in enumerate(kube_hdrs, 1):
                c = ws.cell(row=r, column=ci, value=h)
                c.font      = _font(bold=True, size=9, color="FFFFFF")
                c.fill      = _fill("3D5A80")
                c.alignment = _center()
                c.border    = _border()
            ws.row_dimensions[r].height = 16
            r += 1

            kube_vals = ["Worker", cluster.get("worker_nodes",0),
                         cluster.get("cores_per_node",0),
                         cluster.get("ram_per_node_gb",0),
                         cluster.get("total_ram_gb",0),
                         cluster.get("total_cores",0),
                         cluster.get("storage_gb",0),
                         cluster.get("total_storage_gb",0),
                         "Intel Xeon i7",
                         "AWS Bedrock (Managed)", ""]
            for ci, v in enumerate(kube_vals, 1):
                c = ws.cell(row=r, column=ci, value=v)
                c.font      = _font(size=9)
                c.fill      = _AI_ALT_FILL
                c.alignment = _center()
                c.border    = _border()
            ws.row_dimensions[r].height = 16
            r += 2

            # ── Table 3: Databases (only if present) ───────────────────────
            databases = env_data.get("databases", [])
            if databases:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
                sh3 = ws.cell(row=r, column=1, value="  Databases")
                sh3.font      = _font(bold=True, size=10, color="FFFFFF")
                sh3.fill      = _fill("2E4057")
                sh3.alignment = _left()
                sh3.border    = _border()
                ws.row_dimensions[r].height = 18
                r += 1

                db_hdrs = ["Database", "Type", "Pods", "CPU/Pod", "RAM/Pod (GB)",
                           "Total CPU", "Total RAM (GB)", "Storage (GB)",
                           "Persistent", "", ""]
                for ci, h in enumerate(db_hdrs, 1):
                    c = ws.cell(row=r, column=ci, value=h)
                    c.font      = _font(bold=True, size=9, color="FFFFFF")
                    c.fill      = _fill("3D5A80")
                    c.alignment = _center()
                    c.border    = _border()
                ws.row_dimensions[r].height = 16
                r += 1

                for i, db in enumerate(databases):
                    fill = _AI_ALT_FILL if i % 2 else _AI_WHT_FILL
                    db_vals = [db["name"], db.get("type",""),
                               db.get("pods","—") or "—",
                               db.get("cpu_per_pod","—") or "—",
                               db.get("ram_per_pod","—") or "—",
                               db.get("total_cpu","—") or "—",
                               db.get("total_ram_gb","—") or "—",
                               f"{db['storage_gb']:,} GB",
                               "Yes" if db.get("persistent") else "No", "", ""]
                    for ci, v in enumerate(db_vals, 1):
                        c = ws.cell(row=r, column=ci, value=v)
                        c.font      = _font(size=9)
                        c.fill      = fill
                        c.alignment = _center() if ci > 1 else _left()
                        c.border    = _border()
                    ws.row_dimensions[r].height = 15
                    r += 1
                r += 1

        # GPU / Bedrock note — added to each per-type sheet
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        gn = ws.cell(row=r, column=1,
                     value="⚡ GPU Note: All AI services use AWS Bedrock (Managed LLM Inference API) — "
                           "no self-hosted GPU hardware is required. Bedrock handles LLM inference at scale.")
        gn.font      = _font(size=9, italic=True, color="595959")
        gn.fill      = _fill("F5F5F5")
        gn.alignment = _left()
        gn.border    = _border()
        ws.row_dimensions[r].height = 18

        # Column widths for this sheet
        col_widths = [28, 10, 14, 10, 14, 8, 12, 16, 18, 20, 4]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w


def _build_ai_pricing_sheet(wb: Workbook, ai_sizing: dict, customer: str):
    """
    Builds 'AI Services Pricing' sheet (SaaS only).
    Shows per-environment cost breakdown including:
      - Worker node EC2 cost (estimated from cluster spec)
      - Vector DB (Milvus) — instance cost
      - CRM DB storage
      - AWS Bedrock API
      - 5-year total with inflation
    Uses rough per-node cost estimates based on template instance sizes.
    """
    HOURS_PER_MONTH = 730
    INFLATION_RATE  = 0.04
    # Approximate on-demand hourly rates for 16c/32GB nodes (Intel Xeon)
    # m5.4xlarge: 16 vCPU, 64 GB RAM → ~$0.768/hr
    # r5.2xlarge: 8 vCPU, 64 GB RAM → ~$0.504/hr
    NODE_HOURLY = {
        ("agentic",   "prod"):     0.768,  # 9 × m5.4xlarge
        ("agentic",   "uat"):      0.768,  # 5 × m5.4xlarge
        ("agentic",   "training"): 0.768,
        ("agentic",   "dr_full"):  0.768,
        ("agentic",   "dr_half"):  0.768,
        ("predictive","prod"):     0.384,  # 2 × m5.2xlarge (8c/16GB)
        ("predictive","uat"):      0.384,
        ("predictive","training"): 0.768,
        ("predictive","dr_full"):  0.384,
        ("predictive","dr_half"):  0.384,
        ("genai",     "prod"):     0.768,  # 2 × m5.4xlarge
        ("genai",     "uat"):      0.384,
        ("genai",     "training"): 0.384,
        ("genai",     "dr_full"):  0.768,
        ("genai",     "dr_half"):  0.384,
    }
    EBS_GP3_PER_GB = 0.08

    ws = wb.create_sheet("AI Services Pricing")
    ws.sheet_view.showGridLines = False

    HDR_FILL = _AI_HDR_FILL
    SEC_FILL = _AI_SEC_FILL

    # Title
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value     = f"AI Services — Cloud Pricing Estimate (SaaS) — {customer}"
    t.font      = _font(bold=True, size=13, color="FFFFFF")
    t.fill      = HDR_FILL
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    bedrock_monthly = ai_sizing.get("bedrock_monthly", 3000)

    # Column headers
    hdrs = ["AI Type", "Environment", "Compute (EC2) /mo",
            "Storage (EBS) /mo", "Bedrock API /mo",
            "Total Monthly", "Annual", "5-Year (4% inflation)"]
    r = 2
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font      = _font(bold=True, size=10, color="FFFFFF")
        c.fill      = SEC_FILL
        c.alignment = _center()
        c.border    = _border()
    ws.row_dimensions[r].height = 18
    r += 1

    grand_monthly = 0.0
    environments = ai_sizing.get("environments", {})

    for ai_type, envs_data in environments.items():
        type_label = _AI_TYPE_LABELS.get(ai_type, ai_type.upper())
        type_fill  = _AI_TYPE_FILLS.get(ai_type, SEC_FILL)

        # Section header for this AI type
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        sh = ws.cell(row=r, column=1, value=f"  {type_label}")
        sh.font      = _font(bold=True, size=10, color="FFFFFF")
        sh.fill      = type_fill
        sh.alignment = _left()
        sh.border    = _border()
        ws.row_dimensions[r].height = 18
        r += 1

        for i, (env_key, env_data) in enumerate(envs_data.items()):
            env_label = env_data.get("env_label", env_key)
            cluster   = env_data.get("cluster", {})
            nodes     = cluster.get("worker_nodes", 0)
            stor_gb   = cluster.get("total_storage_gb", 0)

            hourly = NODE_HOURLY.get((ai_type, env_key), 0.384)
            compute_cost  = round(nodes * hourly * HOURS_PER_MONTH, 2)
            storage_cost  = round(stor_gb * EBS_GP3_PER_GB, 2)
            # Bedrock only for agentic + genai; predictive uses Bedrock for scoring API
            bedrock_share = round(bedrock_monthly / max(len(envs_data), 1), 2)
            monthly_total = round(compute_cost + storage_cost + bedrock_share, 2)
            annual        = round(monthly_total * 12, 2)
            five_yr       = round(sum(
                monthly_total * 12 * ((1 + INFLATION_RATE) ** yr)
                for yr in range(1, 6)
            ), 2)
            grand_monthly += monthly_total

            fill = _AI_ALT_FILL if i % 2 else _AI_WHT_FILL
            row_vals = [type_label, env_label, compute_cost,
                        storage_cost, bedrock_share,
                        monthly_total, annual, five_yr]
            for ci, v in enumerate(row_vals, 1):
                c = ws.cell(row=r, column=ci, value=v)
                c.font      = _font(size=9)
                c.fill      = fill
                c.alignment = _center()
                c.border    = _border()
                if isinstance(v, float) and ci >= 3:
                    c.number_format = "$#,##0.00"
            ws.row_dimensions[r].height = 15
            r += 1

    r += 1

    # Grand total row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    tc = ws.cell(row=r, column=1, value="GRAND TOTAL (All AI Types + Environments)")
    tc.font      = _font(bold=True, size=10)
    tc.fill      = TOTAL_FILL
    tc.alignment = _left()
    tc.border    = _border()

    for ci, v in [(6, round(grand_monthly, 2)),
                  (7, round(grand_monthly * 12, 2)),
                  (8, round(sum(grand_monthly * 12 * ((1+INFLATION_RATE)**yr)
                               for yr in range(1, 6)), 2))]:
        c = ws.cell(row=r, column=ci, value=v)
        c.font         = _font(bold=True, size=10)
        c.fill         = TOTAL_FILL
        c.alignment    = _center()
        c.border       = _border()
        c.number_format = "$#,##0.00"
    ws.row_dimensions[r].height = 18
    r += 2

    # Bedrock note
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    nn = ws.cell(row=r, column=1,
                 value=f"⚡ AWS Bedrock API: ${bedrock_monthly:,.0f}/month (user-configured) — "
                       f"split equally across environments. Token-based pricing, no GPU hardware needed.")
    nn.font      = _font(size=9, italic=True, color="595959")
    nn.fill      = _fill("F5F5F5")
    nn.alignment = _left()
    nn.border    = _border()
    ws.row_dimensions[r].height = 18

    # Column widths
    col_widths = [22, 20, 18, 18, 18, 18, 16, 22]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ── Public API ────────────────────────────────────────────────────────────

def generate_excel_reports(
    pricing:      dict,
    distribution: dict,
    metrics:      dict,
    customer:     str   = "Bank-Name",
    output_dir:   str   = "reports",
    env_pricing:  dict  = None,
    db_type:      str   = "PostgreSQL",
    client_mode:  str   = "saas",
    gcp_pricing:  dict  = None,
    comparison:   dict  = None,
    years:        int   = 5,
    include_dr:   bool  = False,
    env_names:    list  = None,
    dr_scale:     float = 1.0,
    ai_sizing:    dict  = None,
    cloud_providers: list = None,   # e.g. ["AWS"], ["GCP"], ["AWS", "GCP"]
) -> dict:
    """
    Generate cloud_sizing.xlsx + aws_pricing.xlsx (AWS) and/or
    gcp_sizing.xlsx + gcp_pricing.xlsx (GCP) based on cloud_providers list.
    For On-Prem clients also generates OpenShift/Kubeadm sizing workbooks.
    Returns dict with file paths.
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    db_selection = pricing.get("db_selection", {}) if pricing else {}

    # ── Normalise cloud providers ──────────────────────────────────────────────
    _providers  = [p.upper() for p in (cloud_providers or ["AWS"])]
    include_aws = "AWS" in _providers
    include_gcp = "GCP" in _providers and gcp_pricing is not None
    _ai = ai_sizing or (distribution or {}).get("ai_nodes", {})
    ch_sizing_data = (distribution or {}).get("clickhouse_nodes", {})

    # ── Workbook 1: Cloud Sizing (AWS) ─────────────────────────────────────────────
    sizing_path = None
    wb1 = Workbook()  # always build; only save if include_aws
    _build_prod_sizing_sheet(wb1, distribution, pricing, metrics, customer)

    # Pre-Prod/DR sizing sheets in Cloud Sizing XLSX only for SaaS
    if client_mode == "saas" and env_pricing:
        preprod = env_pricing.get("preprod_sit_uat")
        dr      = env_pricing.get("dr")
        if preprod:
            pp_names = preprod.get("env_names", [])
            pp_title = " + ".join(pp_names) if pp_names else "Pre-Prod / SIT / UAT"
            _build_env_sizing_sheet(wb1, preprod.get("priced_roles", []), "Pre-Prod Sizing", f"{pp_title} Sizing", customer)
        if dr:
            _build_env_sizing_sheet(wb1, dr.get("priced_roles", []), "DR Sizing", "Disaster Recovery Sizing", customer)
    # (ch_sizing_data extracted above in provider init block)

    # On-Prem: also add DR sizing sheet to cloud_sizing.xlsx for convenience
    if client_mode == "onprem" and include_dr:
        import math as _math
        _mby  = _onprem_metrics_by_year(metrics, years)
        _dr_m = dict(_mby[-1])          # start from Year-N production metrics
        if dr_scale < 1.0:
            # Pilot-light: 1 worker node, infra halved; DB SAN unchanged
            _dr_m["worker_nodes"] = 1
            _dr_m["infra_nodes"]  = max(1, _math.ceil(_dr_m["infra_nodes"] * dr_scale))
            _dr_m["nfs_gb"] = (
                _dr_m["data_gb"] + _dr_m["s3_gb"]
                + _dr_m["worker_nodes"] * 256
                + _dr_m["infra_nodes"]  * 256
                + IMAGE_REGISTRY_GB
            )
            # db_nodes, db_vcpu, db_ram, san_gb — unchanged (full prod DB)
        _build_onprem_env_sheet(
            wb1,
            sheet_name="DR Sizing",
            env_label=f"DR ({int(dr_scale * 100)}% Compute \u2014 Full Prod DB)",
            env_data=_dr_m,
            customer=customer,
            is_prod=True,
            include_reporting_db=True,
            archival_san_gb=5000,
            db_type=db_type if db_type else "SQL Server",
            cluster_name=None,
            ch_sizing=ch_sizing_data if (ch_sizing_data and ch_sizing_data.get("enabled")) else None,
        )

    # ClickHouse sheet in cloud sizing workbook (if enabled)
    if ch_sizing_data and ch_sizing_data.get("enabled") and include_aws:
        _build_clickhouse_sheet(wb1, ch_sizing_data, customer)

    # AI Services sizing sheet (_ai set in provider init block)
    if _ai and _ai.get("enabled") and include_aws:
        _build_ai_sizing_sheet(wb1, _ai, customer)

    if include_aws:
        sizing_path = os.path.join(output_dir, "cloud_sizing.xlsx")
        wb1.save(sizing_path)
        print(f"[excel_exporter] Saved {sizing_path}")

    # ── Workbook 2: Cloud Pricing (AWS, SaaS only) ────────────────────────────
    pricing_path = None
    if client_mode == "saas" and pricing is not None and include_aws:
        wb2 = Workbook()
        _build_pricing_summary_sheet(wb2, pricing, env_pricing, customer)
        _build_prod_pricing_sheet(wb2, pricing, customer)
        if env_pricing:
            preprod = env_pricing.get("preprod_sit_uat")
            dr      = env_pricing.get("dr")
            if preprod:
                pp_names = preprod.get("env_names", [])
                pp_title = " + ".join(pp_names) if pp_names else "SIT + UAT"
                _build_env_pricing_sheet(wb2, preprod, "Pre-Prod Pricing", pp_title, customer, include_forecast=False, row_multiplier=2)
            if dr:
                _build_env_pricing_sheet(wb2, dr, "DR Pricing", "Disaster Recovery", customer, include_forecast=True)
        _build_pupm_sheet(wb2, pricing, env_pricing, metrics, customer, ai_sizing=_ai)
        # ClickHouse sheet in pricing workbook (if enabled)
        ch_s = (distribution or {}).get("clickhouse_nodes", {})
        if ch_s and ch_s.get("enabled"):
            _build_clickhouse_sheet(wb2, ch_s, customer)

        # AI Services pricing sheet in pricing workbook (SaaS only)
        if _ai and _ai.get("enabled"):
            _build_ai_pricing_sheet(wb2, _ai, customer)

        if gcp_pricing:
            _build_gcp_pricing_sheet(wb2, gcp_pricing, customer)
        if comparison:
            _build_comparison_sheet(wb2, comparison, customer)
        pricing_path = os.path.join(output_dir, "cloud_pricing.xlsx")
        wb2.save(pricing_path)
        print(f"[excel_exporter] Saved {pricing_path}")

    # ── Workbook GCP-1: GCP Sizing (SaaS, when GCP selected) ─────────────────
    gcp_sizing_path = None
    if include_gcp and client_mode == "saas":
        wb_gs = Workbook()
        wb_gs.remove(wb_gs.active)          # remove default blank "Sheet"
        prod_gcp_roles = (gcp_pricing or {}).get("priced_roles", [])
        _build_gcp_sizing_sheet(wb_gs, gcp_pricing, customer)
        if env_pricing:
            preprod = env_pricing.get("preprod_sit_uat")
            dr      = env_pricing.get("dr")
            if preprod:
                pp_names = preprod.get("env_names", [])
                pp_title = " + ".join(pp_names) if pp_names else "Pre-Prod / SIT / UAT"
                _build_gcp_env_sizing_sheet(
                    wb_gs, prod_gcp_roles,
                    "Pre-Prod Sizing", f"{pp_title} Sizing", customer,
                    node_scale=0.5,
                )
            if dr:
                _build_gcp_env_sizing_sheet(
                    wb_gs, prod_gcp_roles,
                    "DR Sizing", "Disaster Recovery Sizing", customer,
                    node_scale=0.5,
                )
        gcp_sizing_path = os.path.join(output_dir, "gcp_sizing.xlsx")
        wb_gs.save(gcp_sizing_path)
        print(f"[excel_exporter] Saved {gcp_sizing_path}")

    # ── Workbook GCP-2: GCP Pricing (SaaS, when GCP selected) ────────────────
    gcp_pricing_path = None
    if include_gcp and client_mode == "saas":
        wb_gp = Workbook()
        wb_gp.remove(wb_gp.active)      # remove default blank "Sheet"
        _build_gcp_pricing_sheet(wb_gp, gcp_pricing, customer)
        
        # Inject lower environment pricing for GCP taking scaling sizing
        if env_pricing:
            pp = env_pricing.get("preprod_sit_uat")
            dr = env_pricing.get("dr")
            prod_gcp_roles = (gcp_pricing or {}).get("priced_roles", [])
            
            if pp:
                _build_gcp_env_pricing_sheet(
                    wb_gp, gcp_pricing, prod_gcp_roles,
                    "GCP Pricing (Pre-Prod)", "Pre-Prod / SIT / UAT", customer,
                    node_scale=0.4, include_forecast=False
                )
            if dr:
                _build_gcp_env_pricing_sheet(
                    wb_gp, gcp_pricing, prod_gcp_roles,
                    "GCP Pricing (DR)", "Disaster Recovery (DR)", customer,
                    node_scale=0.5, include_forecast=True
                )

        # AWS vs GCP comparison sheet intentionally omitted from GCP-only workbook
        _build_pupm_sheet(wb_gp, gcp_pricing, env_pricing, metrics, customer,
                          ai_sizing=None, cloud="GCP")
        gcp_pricing_path = os.path.join(output_dir, "gcp_pricing.xlsx")
        wb_gp.save(gcp_pricing_path)
        print(f"[excel_exporter] Saved {gcp_pricing_path}")

    # ── Workbook 3 & 4: On-Prem Sizing Sheets (On-Prem only) ──────────────
    # Both OpenShift and Kubeadm files are ALWAYS generated.
    # Both use the user-selected db_type — same DB, two cluster architectures.
    # DR and Pre-Prod environment sheets are included based on checkbox selections;
    # no pricing is shown — these are purely infrastructure sizing sheets.
    onprem_openshift_path = None
    onprem_kubeadm_path = None
    if client_mode == "onprem":
        primary_db_type = db_type if db_type else "SQL Server"
        db_slug = primary_db_type.lower().replace(" ", "_")

        # File 1: OpenShift cluster — user's selected DB
        onprem_openshift_path = generate_onprem_excel(
            metrics=metrics,
            distribution=distribution,
            customer=customer,
            output_dir=output_dir,
            db_type=primary_db_type,
            years=years,
            filename=f"onprem_openshift_{db_slug}_sizing.xlsx",
            cluster_name="OpenShift",
            include_dr=include_dr,
            env_names=env_names or [],
            dr_scale=dr_scale,
        )

        # File 2: Kubeadm cluster — same user-selected DB
        onprem_kubeadm_path = generate_onprem_excel(
            metrics=metrics,
            distribution=distribution,
            customer=customer,
            output_dir=output_dir,
            db_type=primary_db_type,
            years=years,
            filename=f"onprem_kubeadm_{db_slug}_sizing.xlsx",
            cluster_name="Kubeadm",
            include_dr=include_dr,
            env_names=env_names or [],
            dr_scale=dr_scale,
        )

    # Standalone AI Sizing Workbook (both modes — generated whenever AI is enabled)
    ai_sizing_path = None
    if _ai and _ai.get("enabled"):
        wb_ai = Workbook()
        wb_ai.remove(wb_ai.active)
        _build_ai_sizing_sheet(wb_ai, _ai, customer)
        if client_mode == "saas":
            _build_ai_pricing_sheet(wb_ai, _ai, customer)
        ai_sizing_path = os.path.join(output_dir, "ai_services_sizing_and_pricing.xlsx")
        wb_ai.save(ai_sizing_path)
        print(f"[excel_exporter] Saved AI Services sizing: {ai_sizing_path}")

    return {
        "cloud_sizing":         sizing_path,
        "cloud_pricing":        pricing_path,   # None for on-prem / GCP-only
        "gcp_sizing":           gcp_sizing_path,
        "gcp_pricing":          gcp_pricing_path,
        "onprem_sizing":        onprem_openshift_path,
        "onprem_oracle_sizing": onprem_kubeadm_path,
        "ai_sizing":            ai_sizing_path,
    }


# ── GCP label normalisation ──────────────────────────────────────────────────
_GCP_LABEL_MAP = {
    # Security & Compliance
    "WAF + AWS Shield Advanced":                    "Cloud Armor + Security Command Center",
    "CloudTrail + Config Audit Logs":               "Cloud Audit Logs + Cloud Config",
    # Caching / Session
    "Elasticache Service":                          "Memorystore for Redis",
    "ElastiCache":                                  "Memorystore for Redis",
    # Storage
    "S3 Storage + Replication":                     "Cloud Storage (Standard + Nearline)",
    "S3 Storage":                                   "Cloud Storage (Standard)",
    "Storage: SAN":                                 "Persistent Disk SSD",
    "Storage: SAN (Primary PostgreSQL DB)":         "Persistent Disk SSD (Primary DB)",
    "Storage: SAN (Reporting PostgreSQL DB)":       "Persistent Disk SSD (Reporting DB)",
    "Storage: SAN (Primary SQL Server DB)":         "Persistent Disk SSD (Primary DB)",
    "Storage: SAN (Primary Oracle DB)":             "Persistent Disk SSD (Primary DB)",
    "EBS SAN":                                      "Persistent Disk SSD",
    "Back Up - DB & Infra Logs":                    "Cloud Storage (Nearline — Backup)",
    # Kubernetes / Registry
    "Cloud Managed K8s":                            "GKE Standard Cluster",
    "EKS":                                          "GKE Standard Cluster",
    "Image Registry (ECR)":                         "Artifact Registry",
    "ECR":                                          "Artifact Registry",
    # Networking
    "Load Balancer (Internal + External)":          "Cloud Load Balancing",
    "ALB":                                          "Cloud Load Balancing",
    "NAT Gateway":                                  "Cloud NAT",
    "Bastion Host":                                 "Bastion Host (GCE n2-standard-2)",
    # Observability
    "CloudWatch Monitoring":                        "Cloud Operations Suite",
    "Cloud Operations (Logging + Monitoring)":      "Cloud Operations Suite",
    # Worker nodes — keep label but note GCE
    "Worker Nodes Linux/RHEL for Graphana & Prometheus": "Worker Nodes for Monitoring (Prometheus/Grafana)",
    "Worker Nodes Linux/RHEL for EFK - optional":        "Worker Nodes for Logging (Fluent + Elasticsearch)",
    # AI / ML
    "AWS Bedrock":                                  "Vertex AI (Managed LLM)",
}

def _gcp_label(label: str) -> str:
    """Remap AWS-centric role labels to their GCP equivalents."""
    return _GCP_LABEL_MAP.get(label, label)


# ── GCP Sizing Sheet (Production) ─────────────────────────────────────────────
def _build_gcp_sizing_sheet(wb: Workbook, gcp_pricing: dict, customer: str):
    """GCP Production Sizing sheet — groups by category like the AWS sheet."""
    ws = wb.create_sheet("GCP Prod Sizing")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    NCOLS   = 7
    col_end = get_column_letter(NCOLS)
    _GCP_CAT_MAP = {"S3": "Cloud Storage", "S3 Storage": "Cloud Storage"}

    # Column widths: Category | Role | Nodes | GCP Instance | vCPU | RAM | Storage
    _set_col_widths(ws, [24, 42, 8, 26, 10, 14, 16])

    # ── Title ──────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{col_end}1")
    t = ws["A1"]
    t.value     = f"GCP Cloud Infrastructure Sizing — Production  |  {customer}  |  {datetime.today().strftime('%d %b %Y')}"
    t.font      = _font(bold=True, size=13, color="FFFFFF")
    t.fill      = _fill("1A73E8")   # GCP blue
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    roles = (gcp_pricing or {}).get("priced_roles", [])
    total_nodes = sum(r.get("nodes", 0) or 0 for r in roles)

    ws.merge_cells(f"A2:{col_end}2")
    s = ws["A2"]
    s.value     = f"Total Roles: {len(roles)}   |   Total Nodes: {total_nodes}   |   Google Cloud Platform · Compute Engine"
    s.font      = _font(italic=True, size=10, color="FFFFFF")
    s.fill      = _fill("4285F4")
    s.alignment = _center()
    ws.row_dimensions[2].height = 20

    # ── Headers ────────────────────────────────────────────────────────────
    headers = ["Category", "Service / Role", "Nodes",
               "GCP Instance", "vCPU / Node", "RAM / Node (GB)",
               "Storage / Node (GB)"]
    _header_row(ws, 3, headers)
    ws.row_dimensions[3].height = 28

    # ── Sort by category then cost desc ────────────────────────────────────
    sorted_roles = sorted(roles, key=lambda r: (
        _GCP_CAT_MAP.get(r.get("category", "ZZZ"), r.get("category", "ZZZ")),
        -(r.get("monthly_usd") or 0)
    ))

    row      = 4
    prev_cat = None
    for i, r in enumerate(sorted_roles):
        raw_cat = r.get("category", "Other")
        cat = _GCP_CAT_MAP.get(raw_cat, raw_cat)
        if cat != prev_cat:
            ws.merge_cells(f"A{row}:{col_end}{row}")
            c = ws.cell(row=row, column=1, value=f"  {cat}")
            c.font      = _font(bold=True, size=10, color="FFFFFF")
            c.fill      = _fill("1A73E8")
            c.alignment = _left()
            c.border    = _border()
            ws.row_dimensions[row].height = 20
            row += 1
            prev_cat = cat

        label    = _gcp_label(r.get("label", r.get("role_key", "—")))
        gcp_inst = r.get("gcp_instance_type") or r.get("instance_type") or "—"
        vcpu     = r.get("vcpu_per_node") or "—"
        ram      = r.get("ram_per_node")  or "—"
        stor     = r.get("storage_per_node_gb") or "—"

        vals = [cat, label, r.get("nodes", 0),
                gcp_inst, vcpu, ram, stor]
        _data_row(ws, row, vals, alt=(i % 2 == 0))
        ws.row_dimensions[row].height = 18
        row += 1

    # ── Footer (no prices on a sizing sheet) ───────────────────────────────
    ws.merge_cells(f"A{row}:{col_end}{row}")
    tc = ws.cell(row=row, column=1,
                 value=f"  Total Roles: {len(sorted_roles)}   |   Total Nodes: {total_nodes}")
    tc.font      = _font(italic=True, size=9, color="FFFFFF")
    tc.fill      = _fill("1A73E8")
    tc.alignment = _center()
    ws.row_dimensions[row].height = 18


# ── GCP Env Sizing Sheet (Pre-Prod / DR) ─────────────────────────────────────
def _build_gcp_env_sizing_sheet(
    wb: Workbook,
    prod_gcp_roles: list,
    sheet_name: str,
    title: str,
    customer: str,
    node_scale: float = 0.5,
):
    """
    Build a GCP pre-prod/DR sizing sheet by scaling production GCP roles.
    Uses GCP instance types (not AWS).  node_scale=0.5 → half nodes (min 1).
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    NCOLS   = 7
    col_end = get_column_letter(NCOLS)
    _GCP_CAT_MAP = {"S3": "Cloud Storage", "S3 Storage": "Cloud Storage"}
    _set_col_widths(ws, [24, 42, 8, 26, 10, 14, 16])

    ws.merge_cells(f"A1:{col_end}1")
    t = ws["A1"]
    t.value     = f"{title}  |  {customer}  |  {datetime.today().strftime('%d %b %Y')}"
    t.font      = _font(bold=True, size=13, color="FFFFFF")
    t.fill      = _fill("1A73E8")
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Scale node counts
    scaled = []
    for r in prod_gcp_roles:
        prod_nodes = r.get("nodes") or 1
        env_nodes  = max(1, int(round(prod_nodes * node_scale)))
        scaled.append({**r, "nodes": env_nodes})

    total_nodes = sum(r["nodes"] for r in scaled)

    ws.merge_cells(f"A2:{col_end}2")
    s = ws["A2"]
    s.value     = f"Total Roles: {len(scaled)}   |   Total Nodes: {total_nodes}   |   Google Cloud Platform · Compute Engine"
    s.font      = _font(italic=True, size=10, color="FFFFFF")
    s.fill      = _fill("4285F4")
    s.alignment = _center()
    ws.row_dimensions[2].height = 20

    headers = ["Category", "Service / Role", "Nodes",
               "GCP Instance", "vCPU / Node", "RAM / Node (GB)",
               "Storage / Node (GB)"]
    _header_row(ws, 3, headers)
    ws.row_dimensions[3].height = 28

    sorted_roles = sorted(scaled, key=lambda r: (
        _GCP_CAT_MAP.get(r.get("category", "ZZZ"), r.get("category", "ZZZ")),
        -(r.get("monthly_usd") or 0)
    ))

    row      = 4
    prev_cat = None
    for i, r in enumerate(sorted_roles):
        raw_cat = r.get("category", "Other")
        cat = _GCP_CAT_MAP.get(raw_cat, raw_cat)
        if cat != prev_cat:
            ws.merge_cells(f"A{row}:{col_end}{row}")
            c = ws.cell(row=row, column=1, value=f"  {cat}")
            c.font      = _font(bold=True, size=10, color="FFFFFF")
            c.fill      = _fill("1A73E8")
            c.alignment = _left()
            c.border    = _border()
            ws.row_dimensions[row].height = 20
            row += 1
            prev_cat = cat

        label    = _gcp_label(r.get("label", r.get("role_key", "—")))
        gcp_inst = r.get("gcp_instance_type") or r.get("instance_type") or "—"
        vcpu     = r.get("vcpu_per_node") or "—"
        ram      = r.get("ram_per_node")  or "—"
        stor     = r.get("storage_per_node_gb") or "—"

        vals = [cat, label, r["nodes"],
                gcp_inst, vcpu, ram, stor]
        _data_row(ws, row, vals, alt=(i % 2 == 0))
        ws.row_dimensions[row].height = 18
        row += 1

    # ── Footer ─────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:{col_end}{row}")
    tc = ws.cell(row=row, column=1,
                 value=f"  Total Roles: {len(sorted_roles)}   |   Total Nodes: {total_nodes}")
    tc.font      = _font(italic=True, size=9, color="FFFFFF")
    tc.fill      = _fill("1A73E8")
    tc.alignment = _center()
    ws.row_dimensions[row].height = 18


# ── GCP Pricing Sheet ─────────────────────────────────────────────────────
def _build_gcp_pricing_sheet(wb: Workbook, gcp_pricing: dict, customer: str):
    ws = wb.create_sheet("GCP Pricing")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    # Title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"GCP Compute Engine Pricing — {customer}"
    c.font  = _font(bold=True, size=14, color="FFFFFF")
    c.fill  = _fill("0F9D58")  # GCP green
    c.alignment = _center()
    ws.row_dimensions[1].height = 30

    # Region info
    ws.merge_cells("A2:F2")
    region_label = gcp_pricing.get("region_label", gcp_pricing.get("region",""))
    ws["A2"].value = (
        f"Region: {gcp_pricing.get('region','')}  ({region_label})  ·  "
        f"Generated: {datetime.today().strftime('%d %b %Y')}  ·  "
        f"All On-Demand prices  ·  GCP fallback rates"
    )
    ws["A2"].font      = _font(size=9, color="FFFFFF", italic=True)
    ws["A2"].fill      = _fill("137333")
    ws["A2"].alignment = _center()
    ws.row_dimensions[2].height = 18

    # Table header
    r = 4
    _header_row(ws, r, ["Category", "Role / Service", "Instance Type", "$/hr", "$/month", "$/year"],
                _fill("0F9D58"))
    ws.row_dimensions[r].height = 20
    r += 1

    total_monthly = 0.0
    _GCP_CAT_MAP_P = {"S3": "Cloud Storage", "S3 Storage": "Cloud Storage"}
    sorted_roles = sorted(
        gcp_pricing.get("priced_roles", []),
        key=lambda x: (
            _GCP_CAT_MAP_P.get(x.get("category", "ZZZ"), x.get("category", "ZZZ")),
            -(x.get("monthly_usd") or 0)
        )
    )
    prev_cat = None
    for i, role in enumerate(sorted_roles):
        mo = role.get("monthly_usd", 0) or 0
        total_monthly += mo
        raw_cat = role.get("category", "Other")
        cat = _GCP_CAT_MAP_P.get(raw_cat, raw_cat)
        # Category banner
        if cat != prev_cat:
            ws.merge_cells(f"A{r}:F{r}")
            cb = ws.cell(row=r, column=1, value=f"  {cat}")
            cb.font = _font(bold=True, size=10, color="FFFFFF")
            cb.fill = _fill("0F9D58")
            cb.alignment = _left()
            cb.border = _border()
            ws.row_dimensions[r].height = 20
            r += 1
            prev_cat = cat
        fill = ALT_FILL if i % 2 == 0 else _fill("FFFFFF")
        inst = role.get("gcp_instance_type") or role.get("instance_type") or ""
        vals = [
            cat,
            _gcp_label(role.get("label", "")),
            inst,
            role.get("hourly_usd", 0),
            mo,
            round(mo * 12, 2),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = _font(size=9)
            c.fill      = fill
            c.border    = _border()
            c.alignment = _center() if col >= 4 else _left()
            if col >= 4 and isinstance(val, (int, float)):
                c.number_format = '"$"#,##0.00'
        ws.row_dimensions[r].height = 15
        r += 1

    # Totals
    ws.cell(row=r, column=1, value="TOTAL").font = _font(bold=True, size=10)
    ws.cell(row=r, column=1).fill = TOTAL_FILL
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = _border()
        ws.cell(row=r, column=col).fill   = TOTAL_FILL
    ws.cell(row=r, column=5, value=round(total_monthly, 2)).number_format = '"$"#,##0.00'
    ws.cell(row=r, column=5).font   = _font(bold=True)
    ws.cell(row=r, column=5).fill   = TOTAL_FILL
    ws.cell(row=r, column=6, value=round(total_monthly * 12, 2)).number_format = '"$"#,##0.00'
    ws.cell(row=r, column=6).font   = _font(bold=True)
    ws.cell(row=r, column=6).fill   = TOTAL_FILL
    ws.row_dimensions[r].height = 18
    r += 2

    # 5-Year Forecast
    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"].value     = "5-Year GCP Cost Forecast (4% inflation per year)"
    ws[f"A{r}"].font      = _font(bold=True, size=11, color="FFFFFF")
    ws[f"A{r}"].fill      = _fill("0F9D58")
    ws[f"A{r}"].alignment = _center()
    ws.row_dimensions[r].height = 22
    r += 1

    _header_row(ws, r, ["Year", "Multiplier", "Monthly", "Annual", "Cumulative", "vs Base"],
                _fill("137333"))
    ws.row_dimensions[r].height = 18
    r += 1

    for yr in gcp_pricing.get("inflation_forecast", {}).get("yearly", []):
        fill = ALT_FILL if yr["year"] % 2 == 0 else _fill("FFFFFF")
        vals = [
            f"Year {yr['year']}",
            f"{yr['multiplier']:.4f}×",
            yr["monthly_usd"],
            yr["annual_usd"],
            yr["cumulative_usd"],
            f"+{(yr['multiplier']-1)*100:.1f}%",
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = _font(size=9)
            c.fill      = fill
            c.border    = _border()
            c.alignment = _center()
            if col in (3, 4, 5) and isinstance(val, (int, float)):
                c.number_format = '"$"#,##0.00'
        ws.row_dimensions[r].height = 15
        r += 1


# ── GCP Environment Pricing Sheet (Derived from Sizing) ───────────────────
def _build_gcp_env_pricing_sheet(
    wb: Workbook,
    gcp_pricing: dict,
    prod_roles: list,
    sheet_name: str,
    title: str,
    customer: str,
    node_scale: float = 1.0,
    include_forecast: bool = False,
):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    # Title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"GCP Compute Engine Pricing — {title}  |  {customer}"
    c.font  = _font(bold=True, size=13, color="FFFFFF")
    c.fill  = _fill("0F9D58")
    c.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Subtitle / Region
    ws.merge_cells("A2:F2")
    region_label = gcp_pricing.get("region_label", gcp_pricing.get("region", ""))
    s = ws["A2"]
    s.value = (
        f"Region: {gcp_pricing.get('region','')}  ({region_label})  ·  "
        f"Generated: {datetime.today().strftime('%d %b %Y')}  ·  "
        f"Scaled at {node_scale*100:.0f}% of Prod nodes"
    )
    s.font      = _font(size=9, color="FFFFFF", italic=True)
    s.fill      = _fill("137333")
    s.alignment = _center()
    ws.row_dimensions[2].height = 18

    # Table header
    r = 4
    _header_row(ws, r, ["Category", "Role / Service", "Instance Type", "$/hr", "$/month", "$/year"],
                _fill("0F9D58"))
    ws.row_dimensions[r].height = 20
    r += 1

    total_monthly = 0.0
    _GCP_CAT_MAP_P = {"S3": "Cloud Storage", "S3 Storage": "Cloud Storage"}

    scaled_roles = []
    for role in prod_roles:
        prod_nodes = role.get("nodes") or 1
        env_nodes  = max(1, int(round(prod_nodes * node_scale)))
        
        ratio = env_nodes / prod_nodes if prod_nodes > 0 else 1.0
        env_monthly = (role.get("monthly_usd", 0) or 0) * ratio
        
        scaled_roles.append({
            **role,
            "nodes": env_nodes,
            "monthly_usd": env_monthly
        })

    sorted_roles = sorted(
        scaled_roles,
        key=lambda x: (
            _GCP_CAT_MAP_P.get(x.get("category", "ZZZ"), x.get("category", "ZZZ")),
            -(x.get("monthly_usd") or 0)
        )
    )

    prev_cat = None
    for i, role in enumerate(sorted_roles):
        mo = role.get("monthly_usd", 0) or 0
        total_monthly += mo
        raw_cat = role.get("category", "Other")
        cat = _GCP_CAT_MAP_P.get(raw_cat, raw_cat)
        # Category banner
        if cat != prev_cat:
            ws.merge_cells(f"A{r}:F{r}")
            cb = ws.cell(row=r, column=1, value=f"  {cat}")
            cb.font = _font(bold=True, size=10, color="FFFFFF")
            cb.fill = _fill("0F9D58")
            cb.alignment = _left()
            cb.border = _border()
            ws.row_dimensions[r].height = 20
            r += 1
            prev_cat = cat
        
        fill = ALT_FILL if i % 2 == 0 else _fill("FFFFFF")
        inst = role.get("gcp_instance_type") or role.get("instance_type") or ""
        
        vals = [
            cat,
            f"{role.get('nodes')}× " + _gcp_label(role.get("label", "")),
            inst,
            role.get("hourly_usd", 0),
            mo,
            round(mo * 12, 2),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = _font(size=9)
            c.fill      = fill
            c.border    = _border()
            c.alignment = _center() if col >= 4 else _left()
            if col >= 4 and isinstance(val, (int, float)):
                c.number_format = '"$"#,##0.00'
        ws.row_dimensions[r].height = 15
        r += 1

    # Totals
    ws.cell(row=r, column=1, value="TOTAL").font = _font(bold=True, size=10)
    ws.cell(row=r, column=1).fill = TOTAL_FILL
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = _border()
        ws.cell(row=r, column=col).fill   = TOTAL_FILL
    ws.cell(row=r, column=5, value=round(total_monthly, 2)).number_format = '"$"#,##0.00'
    ws.cell(row=r, column=5).font   = _font(bold=True)
    ws.cell(row=r, column=5).fill   = TOTAL_FILL
    ws.cell(row=r, column=6, value=round(total_monthly * 12, 2)).number_format = '"$"#,##0.00'
    ws.cell(row=r, column=6).font   = _font(bold=True)
    ws.cell(row=r, column=6).fill   = TOTAL_FILL
    ws.row_dimensions[r].height = 18
    r += 2

    # 5-Year Forecast
    if include_forecast:
        ws.merge_cells(f"A{r}:F{r}")
        ws[f"A{r}"].value     = "5-Year GCP Cost Forecast (4% inflation per year)"
        ws[f"A{r}"].font      = _font(bold=True, size=11, color="FFFFFF")
        ws[f"A{r}"].fill      = _fill("0F9D58")
        ws[f"A{r}"].alignment = _left()
        ws.row_dimensions[r].height = 22
        r += 1

        _header_row(ws, r, ["Year", "Multiplier", "Monthly", "Annual", "Cumulative", "vs Base"],
                    _fill("137333"))
        ws.row_dimensions[r].height = 18
        r += 1

        cumulative = 0.0
        for yr_idx in range(1, 6):
            mult = (1 + 0.04) ** yr_idx
            annual = total_monthly * 12 * mult
            cumulative += annual

            fill = ALT_FILL if yr_idx % 2 == 0 else _fill("FFFFFF")
            vals = [
                f"Year {yr_idx}",
                f"{mult:.4f}×",
                total_monthly * mult,
                annual,
                cumulative,
                f"+{(mult - 1)*100:.1f}%",
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.font      = _font(size=9)
                c.fill      = fill
                c.border    = _border()
                c.alignment = _center()
                if col in (3, 4, 5) and isinstance(val, (int, float)):
                    c.number_format = '"$"#,##0.00'
            ws.row_dimensions[r].height = 15
        r += 1


# ── Comparison Sheet ──────────────────────────────────────────────────────
def _build_comparison_sheet(wb: Workbook, comparison: dict, customer: str):
    ws = wb.create_sheet("AWS vs GCP Comparison")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [22, 18, 18, 18, 18, 18, 14]):
        ws.column_dimensions[col].width = w

    s = comparison.get("summary", {})

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"].value     = f"AWS vs GCP Cost Comparison — {customer}"
    ws["A1"].font      = _font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = _fill("1F4E79")
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 30

    # Summary sub-header
    ws.merge_cells("A2:G2")
    aws_reg = s.get("aws_region", "us-east-1")
    gcp_reg = s.get("gcp_region", "us-central1")
    ws["A2"].value = (
        f"AWS Region: {aws_reg}   |   GCP Region: {gcp_reg}   |   "
        f"Generated: {datetime.today().strftime('%d %b %Y')}   |   On-Demand pricing"
    )
    ws["A2"].font      = _font(size=9, color="FFFFFF", italic=True)
    ws["A2"].fill      = _fill("2E75B6")
    ws["A2"].alignment = _center()
    ws.row_dimensions[2].height = 18

    # KPI summary row
    r = 4
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"].value     = "Cost Summary"
    ws[f"A{r}"].font      = _font(bold=True, size=11, color="FFFFFF")
    ws[f"A{r}"].fill      = SECTION_FILL
    ws[f"A{r}"].alignment = _center()
    ws.row_dimensions[r].height = 20
    r += 1

    summary_rows = [
        ("Monthly Cost (AWS)",   s.get("aws_monthly",  0), "$#,##0.00"),
        ("Monthly Cost (GCP)",   s.get("gcp_monthly",  0), "$#,##0.00"),
        ("Monthly Difference",   s.get("diff_monthly", 0), "$#,##0.00"),
        ("Cheaper Cloud (Monthly)", s.get("cheaper_monthly", "AWS"), "@"),
        ("Annual Cost (AWS)",    s.get("aws_annual",   0), "$#,##0.00"),
        ("Annual Cost (GCP)",    s.get("gcp_annual",   0), "$#,##0.00"),
        ("5-Year Total (AWS)",   s.get("aws_5year",    0), "$#,##0.00"),
        ("5-Year Total (GCP)",   s.get("gcp_5year",    0), "$#,##0.00"),
        ("5-Year Savings",       s.get("diff_5year",   0), "$#,##0.00"),
        ("5-Year Winner",        s.get("cheaper_5year","AWS"), "@"),
    ]
    for label, val, fmt in summary_rows:
        fill = GREEN_FILL if label in ("Cheaper Cloud (Monthly)","5-Year Winner") else _fill("FFFFFF")
        ws.cell(row=r, column=1, value=label).font      = _font(bold=True, size=10)
        ws.cell(row=r, column=1).alignment  = _left()
        ws.cell(row=r, column=1).border     = _border()
        ws.cell(row=r, column=2, value=val).font        = _font(size=10)
        ws.cell(row=r, column=2).alignment  = _center()
        ws.cell(row=r, column=2).border     = _border()
        ws.cell(row=r, column=2).fill       = fill
        if fmt != "@" and isinstance(val, (int, float)):
            ws.cell(row=r, column=2).number_format = fmt
        for col in range(1, 3):
            ws.cell(row=r, column=col).fill = fill
        ws.row_dimensions[r].height = 15
        r += 1
    r += 1

    # Category comparison table
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"].value     = "Category-Level Cost Comparison"
    ws[f"A{r}"].font      = _font(bold=True, size=11, color="FFFFFF")
    ws[f"A{r}"].fill      = SECTION_FILL
    ws[f"A{r}"].alignment = _center()
    ws.row_dimensions[r].height = 20
    r += 1

    _header_row(ws, r, ["Category", "AWS $/mo", "GCP $/mo", "Difference", "% Diff", "Cheaper", ""],
                _fill("1F4E79"))
    ws.row_dimensions[r].height = 18
    r += 1

    for i, row in enumerate(comparison.get("category_comparison", [])):
        fill = ALT_FILL if i % 2 == 0 else _fill("FFFFFF")
        diff = row["diff"]
        vals = [
            row["category"],
            row["aws_monthly"],
            row["gcp_monthly"],
            abs(diff),
            f"{abs(row.get('pct_diff',0)):.1f}%",
            row["cheaper"],
            "✔",
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = _font(size=9, bold=(col==7))
            c.fill      = GREEN_FILL if col == 6 else fill
            c.border    = _border()
            c.alignment = _center() if col > 1 else _left()
            if col in (2, 3, 4) and isinstance(val, (int, float)):
                c.number_format = '"$"#,##0.00'
        ws.row_dimensions[r].height = 15
        r += 1
    r += 1

    # Year-by-year comparison
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"].value     = "5-Year Year-by-Year Comparison"
    ws[f"A{r}"].font      = _font(bold=True, size=11, color="FFFFFF")
    ws[f"A{r}"].fill      = SECTION_FILL
    ws[f"A{r}"].alignment = _center()
    ws.row_dimensions[r].height = 20
    r += 1

    _header_row(ws, r, ["Year", "AWS Monthly", "GCP Monthly", "AWS Annual", "GCP Annual",
                         "AWS Cumul.", "GCP Cumul."], _fill("1F4E79"))
    ws.row_dimensions[r].height = 18
    r += 1

    for i, yr in enumerate(comparison.get("yearly_comparison", [])):
        fill = ALT_FILL if i % 2 == 0 else _fill("FFFFFF")
        cheaper_fill = GREEN_FILL if yr.get("cheaper") == "AWS" else _fill("E8F5E9")
        vals = [
            f"Year {yr['year']}",
            yr["aws_monthly"], yr["gcp_monthly"],
            yr["aws_annual"],  yr["gcp_annual"],
            yr["aws_cumulative"], yr["gcp_cumulative"],
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = _font(size=9)
            c.fill      = fill
            c.border    = _border()
            c.alignment = _center()
            if col > 1 and isinstance(val, (int, float)):
                c.number_format = '"$"#,##0.00'
        # Highlight cheaper cloud columns
        aws_col = 2 if yr.get("cheaper") == "AWS" else None
        gcp_col = 3 if yr.get("cheaper") == "GCP" else None
        if aws_col: ws.cell(row=r, column=aws_col).fill = GREEN_FILL
        if gcp_col: ws.cell(row=r, column=gcp_col).fill = GREEN_FILL
        ws.row_dimensions[r].height = 15


# ══════════════════════════════════════════════════════════════════════════════
# On-Premise OpenShift / Kubeadm Sizing Workbook
# ══════════════════════════════════════════════════════════════════════════════

MASTER_NODES = 3          # Always constant per user requirement
BOOTSTRAP_CPU, BOOTSTRAP_RAM, BOOTSTRAP_STORAGE = 2, 16, 256
MASTER_CPU,    MASTER_RAM,    MASTER_STORAGE    = 4, 16, 256
WORKER_CPU,    WORKER_RAM,    WORKER_STORAGE    = 8, 32, 256
INFRA_CPU,     INFRA_RAM,     INFRA_STORAGE     = 4, 32, 256
BASTION2_CPU,  BASTION2_RAM,  BASTION2_STORAGE  = 2, 8,  256
DB_WIN_STORAGE = 300      # per Windows node
IMAGE_REGISTRY_GB = 1024


def _onprem_metrics_by_year(metrics: dict, years: int) -> list:
    """
    Project per-year metrics for the On-Prem sizing sheets.
    Worker nodes grow ~5%/yr, data 10%/yr (matching template ratios 7→8 over 4 yrs).
    Returns list of dicts, one per year (index 0 = year 1).
    """
    base_workers  = int(metrics.get("total_workernodes", 7))
    base_data_gb  = int(metrics.get("data_size_gb",  14000))
    base_s3_gb    = int(metrics.get("s3_size_gb",     7100))
    base_vcpus    = int(metrics.get("total_vcpus_workernode", 56))

    # Infer DB node sizing from total_vcpus (template uses memory-intensive nodes)
    # Template Y1: 2 Windows nodes × 18 cores = 36 vCPU ≈ 64% of worker vCPUs
    base_db_vcpu   = max(18, int(base_vcpus * 0.64))
    base_db_ram    = base_db_vcpu * 16   # 16 GB per vCPU
    base_db_nodes  = 2                   # always 2 primary DB nodes in PROD

    result = []
    for y in range(1, years + 1):
        # Worker nodes: ceil of 5% growth
        worker_y = max(base_workers, int(base_workers * (1.05 ** (y - 1))))
        # Infra nodes: 3 for PROD (stable)
        infra_y  = 3
        # Data size: grows 10% per year
        data_y   = int(base_data_gb * (1.10 ** (y - 1)))
        s3_y     = int(base_s3_gb   * (1.10 ** (y - 1)))
        # DB vCPUs scale with data size
        scaler    = data_y / max(base_data_gb, 1)
        db_vcpu_y = max(base_db_vcpu, int(base_db_vcpu * scaler))
        db_ram_y  = db_vcpu_y * 16
        # NFS storage = data + S3 + per-node overhead + image registry
        nfs_y = data_y + s3_y + worker_y * 256 + infra_y * 256 + IMAGE_REGISTRY_GB
        # SAN mirrors data size
        san_y = data_y
        result.append({
            "year":         y,
            "worker_nodes": worker_y,
            "infra_nodes":  infra_y,
            "data_gb":      data_y,
            "s3_gb":        s3_y,
            "db_nodes":     base_db_nodes,
            "db_vcpu":      db_vcpu_y,
            "db_ram":       db_ram_y,
            "nfs_gb":       nfs_y,
            "san_gb":       san_y,
            "app_servers":  worker_y,
        })
    return result


def _build_onprem_data_sheet(wb: Workbook, metrics_by_year: list, customer: str):
    """Builds the 'Data' summary sheet matching the template's Data tab."""
    ws = wb.create_sheet("Data")
    ws.sheet_view.showGridLines = False

    NUM_YEARS   = len(metrics_by_year)
    yr_suffixes = ["1st", "2nd", "3rd"] + [f"{y}th" for y in range(4, NUM_YEARS + 1)]
    year_labels = yr_suffixes[:NUM_YEARS]

    HDR_FILL = _fill("1F4E79")
    SEC_FILL = _fill("2E75B6")
    ALT_FILL = _fill("EBF3FB")
    WHT_FILL = _fill("FFFFFF")

    last_col    = 2 + NUM_YEARS
    last_letter = get_column_letter(last_col)

    # Title
    ws.merge_cells(f"A1:{last_letter}1")
    t = ws["A1"]
    t.value     = f"On-Premise Infrastructure Metrics — {customer} — {datetime.today().strftime('%d %b %Y')}"
    t.font      = _font(bold=True, size=13, color="FFFFFF")
    t.fill      = HDR_FILL
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Header row: blank | "Year >" | 1st | 2nd | …
    ws.cell(row=2, column=1, value="Metric").font = _font(bold=True, color="FFFFFF", size=10)
    ws.cell(row=2, column=1).fill = HDR_FILL
    ws.cell(row=2, column=1).alignment = _left()
    ws.cell(row=2, column=1).border = _border()
    ws.cell(row=2, column=2, value="Year >").font = _font(bold=True, color="FFFFFF", size=10)
    ws.cell(row=2, column=2).fill = HDR_FILL
    ws.cell(row=2, column=2).alignment = _center()
    ws.cell(row=2, column=2).border = _border()
    for i, label in enumerate(year_labels):
        c = ws.cell(row=2, column=3+i, value=label)
        c.font      = _font(bold=True, color="FFFFFF", size=10)
        c.fill      = SEC_FILL
        c.alignment = _center()
        c.border    = _border()
    ws.row_dimensions[2].height = 22

    def _sec(row, title):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
        c = ws.cell(row=row, column=1, value=f"  {title}")
        c.font      = _font(bold=True, color="FFFFFF", size=10)
        c.fill      = SEC_FILL
        c.alignment = _left()
        c.border    = _border()
        ws.row_dimensions[row].height = 18

    def _dat(row, label, values, alt=False):
        fill = ALT_FILL if alt else WHT_FILL
        c = ws.cell(row=row, column=1, value=label)
        c.font = _font(size=10); c.fill = fill; c.alignment = _left(); c.border = _border()
        ws.cell(row=row, column=2).fill   = fill
        ws.cell(row=row, column=2).border = _border()
        for i, v in enumerate(values):
            cell = ws.cell(row=row, column=3+i, value=v)
            cell.font = _font(size=10); cell.fill = fill
            cell.alignment = _center(); cell.border = _border()
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0"
        ws.row_dimensions[row].height = 16

    r = 3
    _sec(r, "CRMNEXT App Servers"); r += 1
    _dat(r, "Total App Servers (8 Phy. Cores/16vCPU, 32GB RAM Each)",
         [m["app_servers"] for m in metrics_by_year]); r += 1

    _sec(r, "CRMNEXT Database Server"); r += 1
    _dat(r, "Total Cores Required",
         [m["db_vcpu"] // 2 for m in metrics_by_year], alt=True); r += 1
    _dat(r, "Total vCPUs Required",
         [m["db_vcpu"] for m in metrics_by_year]); r += 1

    _sec(r, "Volumes"); r += 1
    _dat(r, "Data size (GB)",
         [m["data_gb"] for m in metrics_by_year], alt=True); r += 1
    _dat(r, "S3 size (GB)",
         [m["s3_gb"] for m in metrics_by_year]); r += 1

    _set_col_widths(ws, [52, 12] + [14] * NUM_YEARS)


def _build_onprem_env_sheet(
    wb: Workbook,
    sheet_name: str,
    env_label: str,
    env_data: dict,
    customer: str,
    is_prod: bool = False,
    include_reporting_db: bool = False,
    archival_san_gb: float = 0,
    db_type: str = "SQL Server",
    cluster_name: str = None,   # Override cluster label: "OpenShift" or "Kubeadm"
    ch_sizing: dict = None,     # ClickHouse sizing dict from distribution["clickhouse_nodes"]
):
    """
    Builds one environment sheet (PROD-XYr, DR, PRE-PROD, UAT, SIT)
    matching the standard on-prem sizing layout exactly.
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    worker_nodes = int(env_data["worker_nodes"])
    infra_nodes  = int(env_data["infra_nodes"])
    db_nodes     = int(env_data["db_nodes"])
    db_vcpu      = int(env_data["db_vcpu"])
    db_ram       = int(env_data["db_ram"])
    data_gb      = int(env_data["data_gb"])
    s3_gb        = int(env_data["s3_gb"])
    nfs_gb       = int(env_data["nfs_gb"])
    san_gb       = int(env_data["san_gb"])

    # Use explicit override if provided; otherwise derive from db_type
    if cluster_name is None:
        cluster_name = "Kubeadm" if db_type == "Oracle" else "OpenShift"

    HDR_FILL  = _fill("1F4E79")
    SEC_FILL  = _fill("2E75B6")
    ALT_FILL  = _fill("EBF3FB")
    WHT_FILL  = _fill("FFFFFF")

    # ── Title row ────────────────────────────────────────────────────────────
    ws.merge_cells("B1:N1")
    t = ws["B1"]
    t.value     = f"{env_label} Environment On Premise"
    t.font      = _font(bold=True, size=12, color="FFFFFF")
    t.fill      = HDR_FILL
    t.alignment = _center()
    ws.row_dimensions[1].height = 26

    # ── Column header row ────────────────────────────────────────────────────
    col_headers = [
        "S.N.", "Services/Instances", "Nodes", "Instance Type", "Remarks",
        "Per Node\nCPU cores", "Per Node\nRAM", "Per Node\nStorage (GB)",
        "Total\nCPU cores", "Total\nRAM\n(GB)", "Total\nStorage\n(TB)",
        None, "Remarks", None,
    ]
    for i, h in enumerate(col_headers):
        c = ws.cell(row=2, column=2+i, value=h)
        c.font      = _font(bold=True, size=9, color="FFFFFF")
        c.fill      = HDR_FILL
        c.alignment = _center()
        c.border    = _border()
    ws.row_dimensions[2].height = 36

    def _sec(row, sn, title):
        ws.cell(row=row, column=2, value=sn).font      = _font(bold=True, size=10, color="FFFFFF")
        ws.cell(row=row, column=2).fill      = SEC_FILL
        ws.cell(row=row, column=2).border    = _border()
        ws.cell(row=row, column=2).alignment = _center()
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=15)
        c = ws.cell(row=row, column=3, value=title)
        c.font = _font(bold=True, size=10, color="FFFFFF")
        c.fill = SEC_FILL; c.alignment = _left(); c.border = _border()
        for col in [13, 14, 15]:
            ws.cell(row=row, column=col).fill   = SEC_FILL
            ws.cell(row=row, column=col).border = _border()
        ws.row_dimensions[row].height = 20

    def _dr(row, sn, service, nodes, inst_type, remarks,
            cpu_n, ram_n, stor_n, tot_cpu, tot_ram, tot_stor_tb,
            remarks2="", alt=False):
        fill = ALT_FILL if alt else WHT_FILL
        vals = [sn, service, nodes, inst_type, remarks,
                cpu_n, ram_n, stor_n, tot_cpu, tot_ram, tot_stor_tb, None, remarks2, None]
        for i, v in enumerate(vals):
            col = 2 + i
            c = ws.cell(row=row, column=col, value=v)
            c.font      = _font(size=9)
            c.fill      = fill
            c.alignment = _center() if col > 3 else _left()
            c.border    = _border()
            if col == 12 and isinstance(v, float):
                c.number_format = "0.00####"
        ws.row_dimensions[row].height = 16

    def _tb(n, gb):
        return round(n * gb / 1024, 8)

    r = 3

    # Section 1: Cluster
    _sec(r, 1, f"{cluster_name} Cluster"); r += 1
    _dr(r, None, "Bootstrap machine", 1, "Memory Intensive",
        f"{cluster_name} vendor to confirm on sizing.",
        BOOTSTRAP_CPU, BOOTSTRAP_RAM, BOOTSTRAP_STORAGE,
        BOOTSTRAP_CPU, BOOTSTRAP_RAM, _tb(1, BOOTSTRAP_STORAGE), alt=True); r += 1
    _dr(r, None, "Bastion Host", 1, "Memory Intensive", "",
        BOOTSTRAP_CPU, BOOTSTRAP_RAM, BOOTSTRAP_STORAGE,
        BOOTSTRAP_CPU, BOOTSTRAP_RAM, _tb(1, BOOTSTRAP_STORAGE)); r += 1
    _dr(r, None, "Master Nodes", MASTER_NODES, "Compute Intensive", "",
        MASTER_CPU, MASTER_RAM, MASTER_STORAGE,
        MASTER_NODES*MASTER_CPU, MASTER_NODES*MASTER_RAM,
        _tb(MASTER_NODES, MASTER_STORAGE), alt=True); r += 1
    _dr(r, None, "Linux Worker Nodes (BUSINESSNEXT)", worker_nodes, "Compute Intensive", "",
        WORKER_CPU, WORKER_RAM, WORKER_STORAGE,
        worker_nodes*WORKER_CPU, worker_nodes*WORKER_RAM,
        _tb(worker_nodes, WORKER_STORAGE)); r += 1

    # Section 2: Operational Services
    _sec(r, 2, "Operational Services"); r += 1
    _dr(r, None, "Infra Nodes (Grafana & Prometheus, EFK, Redis)",
        infra_nodes, "Compute Intensive", "",
        INFRA_CPU, INFRA_RAM, INFRA_STORAGE,
        infra_nodes*INFRA_CPU, infra_nodes*INFRA_RAM,
        _tb(infra_nodes, INFRA_STORAGE), alt=True); r += 1
    _dr(r, None, "NFS Storage", 1, "", "",
        None, None, nfs_gb, None, None, round(nfs_gb/1024, 8)); r += 1

    sn = 3

    vcpu_per = db_vcpu // max(db_nodes, 1)
    ram_per  = db_ram  // max(db_nodes, 1)

    if db_type == "Oracle":
        # Section 3: Oracle Database
        _sec(r, sn, "Oracle Database"); r += 1
        _dr(r, None, "OEL/RHEL", db_nodes, "Memory Intensive", "",
            vcpu_per, ram_per, DB_WIN_STORAGE,
            db_vcpu, db_ram, _tb(db_nodes, DB_WIN_STORAGE), alt=True); r += 1
        _dr(r, None, "Oracle License", db_nodes, "Memory Intensive", "",
            None, None, None, None, None, None); r += 1
        _dr(r, None, "Storage: SAN", db_nodes, "20K IOPS", "",
            None, None, san_gb, None, None,
            _tb(db_nodes, san_gb), alt=True); r += 1
        if is_prod:
            _dr(r, None, "RAC - Active/Active", 1, "20K IOPS",
                "As per current archival database size",
                None, None, archival_san_gb, None, None,
                _tb(1, archival_san_gb)); r += 1
            _dr(r, None, "Dataguard", 1, "", "",
                None, None, None, None, None, None, alt=True); r += 1
        sn += 1

        # Section 4: Oracle Database - Reporting (PROD only)
        if is_prod and include_reporting_db:
            _sec(r, sn, "Oracle Database - Reporting"); r += 1
            _dr(r, None, "OEL/RHEL", 1, "Memory Intensive", "",
                vcpu_per, ram_per, DB_WIN_STORAGE,
                vcpu_per, ram_per, _tb(1, DB_WIN_STORAGE), alt=True); r += 1
            _dr(r, None, "Oracle License", 1, "Memory Intensive", "",
                None, None, None, None, None, None); r += 1
            _dr(r, None, "Storage: SAN", 1, "20K IOPS", "",
                None, None, san_gb, None, None,
                _tb(1, san_gb), alt=True); r += 1
            sn += 1
    elif db_type == "SQL Server":
        # Section 3: MSSQL Enterprise Database
        _sec(r, sn, "MSSQL Enterprise Database"); r += 1
        _dr(r, None, "Windows Nodes", db_nodes, "Memory Intensive", "",
            vcpu_per, ram_per, DB_WIN_STORAGE,
            db_vcpu, db_ram, _tb(db_nodes, DB_WIN_STORAGE), alt=True); r += 1
        _dr(r, None, "SQL License", db_nodes, "Memory Intensive", "",
            None, None, None, None, None, None); r += 1
        _dr(r, None, "Storage: SAN", db_nodes, "20K IOPS", "",
            None, None, san_gb, None, None,
            _tb(db_nodes, san_gb), alt=True); r += 1
        if is_prod:
            _dr(r, None, "Storage: SAN (Archival)", db_nodes, "20K IOPS",
                "As per current archival database size",
                None, None, archival_san_gb, None, None,
                _tb(db_nodes, archival_san_gb)); r += 1
        sn += 1

        # Section 4: MSSQL Reporting DB (PROD only)
        if is_prod and include_reporting_db:
            _sec(r, sn, "MSSQL Enterprise Database - Reporting"); r += 1
            _dr(r, None, "Windows Nodes", 1, "Memory Intensive", "",
                vcpu_per, ram_per, DB_WIN_STORAGE,
                vcpu_per, ram_per, _tb(1, DB_WIN_STORAGE), alt=True); r += 1
            _dr(r, None, "SQL License", 1, "Memory Intensive", "",
                None, None, None, None, None, None); r += 1
            _dr(r, None, "Storage: SAN", 1, "20K IOPS", "",
                None, None, san_gb, None, None,
                _tb(1, san_gb), alt=True); r += 1
            sn += 1
    else:
        # Section 3: PostgreSQL Database (Patroni HA)
        _sec(r, sn, "PostgreSQL Database (Patroni HA)"); r += 1
        _dr(r, None, "Linux Nodes — Database server (Primary)", db_nodes, "Memory Intensive", "",
            vcpu_per, ram_per, DB_WIN_STORAGE,
            db_vcpu, db_ram, _tb(db_nodes, DB_WIN_STORAGE), alt=True); r += 1
        _dr(r, None, "Nodes — Cluster (etcd+haproxy, pgbackrest)", 4, "Memory Intensive", "",
            2, 8, 100, 8, 32, _tb(4, 100)); r += 1
        _dr(r, None, "Storage: SAN (Primary PostgreSQL DB)", db_nodes, "10K IOPS", "",
            None, None, san_gb, None, None,
            _tb(db_nodes, san_gb), alt=True); r += 1
        if is_prod:
            _dr(r, None, "Storage: SAN (Archival)", db_nodes, "10K IOPS",
                "As per current archival database size",
                None, None, archival_san_gb, None, None,
                _tb(db_nodes, archival_san_gb)); r += 1
        sn += 1

        # Section 4: PostgreSQL Reporting DB (PROD only)
        if is_prod and include_reporting_db:
            _sec(r, sn, "PostgreSQL Database - Reporting"); r += 1
            _dr(r, None, "Linux Nodes — Reporting replica", 1, "Memory Intensive", "",
                vcpu_per, ram_per, DB_WIN_STORAGE,
                vcpu_per, ram_per, _tb(1, DB_WIN_STORAGE), alt=True); r += 1
            _dr(r, None, "Storage: SAN", 1, "10K IOPS", "",
                None, None, san_gb, None, None,
                _tb(1, san_gb), alt=True); r += 1
            sn += 1

    # ── ClickHouse OLAP section (if enabled) ────────────────────────────────
    if ch_sizing and ch_sizing.get("enabled"):
        db_cl  = ch_sizing.get("db_cluster", {})
        kp_cl  = ch_sizing.get("keeper_cluster", {})
        ch_sum = ch_sizing.get("summary", {})

        if is_prod:
            # Full cluster for PROD/DR environments
            ch_db_nodes    = db_cl.get("total_nodes", 2)
            ch_db_vcpu     = db_cl.get("vcpu_per_node", 8)
            ch_db_ram      = db_cl.get("ram_per_node", 32)
            ch_db_stor     = db_cl.get("storage_per_node_gb", 5000)
            ch_kp_nodes    = kp_cl.get("total_nodes", 3)
            ch_kp_vcpu     = kp_cl.get("vcpu_per_node", 4)
            ch_kp_ram      = kp_cl.get("ram_per_node", 16)
            ch_kp_stor     = kp_cl.get("storage_per_node_gb", 200)
            ch_shards      = db_cl.get("num_shards", 1)
            ch_replicas    = db_cl.get("replicas_per_shard", 2)
        else:
            # Minimal cluster for Pre-Prod/UAT/SIT
            ch_db_nodes    = 1
            ch_db_vcpu     = db_cl.get("vcpu_per_node", 8)
            ch_db_ram      = db_cl.get("ram_per_node", 32)
            ch_db_stor     = min(db_cl.get("storage_per_node_gb", 5000), 2000)
            ch_kp_nodes    = 1
            ch_kp_vcpu     = kp_cl.get("vcpu_per_node", 4)
            ch_kp_ram      = kp_cl.get("ram_per_node", 16)
            ch_kp_stor     = kp_cl.get("storage_per_node_gb", 200)
            ch_shards      = 1
            ch_replicas    = 1

        topology = f"{ch_shards} shard(s) × {ch_replicas} replica(s)"

        _sec(r, sn, f"ClickHouse OLAP Database — {topology}"); r += 1
        _dr(r, None, "ClickHouse DB Nodes (Self-Hosted)",
            ch_db_nodes, "Memory Intensive",
            f"Columnar OLAP — {topology}",
            ch_db_vcpu, ch_db_ram, ch_db_stor,
            ch_db_nodes * ch_db_vcpu, ch_db_nodes * ch_db_ram,
            _tb(ch_db_nodes, ch_db_stor), alt=True); r += 1
        _dr(r, None, "ClickHouse DB — SSD/SAN Storage",
            ch_db_nodes, "High IOPS (gp3 equiv)",
            "Per-node SSD for columnar merge-tree",
            None, None, ch_db_stor, None, None,
            _tb(ch_db_nodes, ch_db_stor)); r += 1
        _dr(r, None, f"ClickHouse Keeper Nodes (Raft Quorum)",
            ch_kp_nodes, "General Purpose",
            f"{ch_kp_nodes}-node Raft quorum for metadata coordination",
            ch_kp_vcpu, ch_kp_ram, ch_kp_stor,
            ch_kp_nodes * ch_kp_vcpu, ch_kp_nodes * ch_kp_ram,
            _tb(ch_kp_nodes, ch_kp_stor), alt=True); r += 1
        sn += 1

    # S3 section
    _sec(r, sn, "S3"); r += 1
    _dr(r, None, "S3 Storage", 1,
        "To be consumed from main NFS storage",
        "Will be based on the number of documents provided, "
        "Considering per document size per Service Request of 512KB Each.",
        None, None, s3_gb, None, None, round(s3_gb/1024, 8), alt=True); r += 1
    sn += 1

    # Infrastructure & Monitoring
    _sec(r, sn, "Infrastructure & Monitoring"); r += 1
    _dr(r, None, "SSL", 1, "Provided by Bank", "",
        None, None, None, None, None, None); r += 1
    _dr(r, None, "Image Registry", 1, "For Containers docker images", "",
        None, None, IMAGE_REGISTRY_GB, None, None,
        round(IMAGE_REGISTRY_GB/1024, 8), alt=True); r += 1
    _dr(r, None, "Web Application Firewall (WAF)", 1, "",
        "As per banks security policy.",
        None, None, None, None, None, None); r += 1
    _dr(r, None, "Bastion Host", 1, "", "As per banks policy.",
        BASTION2_CPU, BASTION2_RAM, BASTION2_STORAGE,
        BASTION2_CPU, BASTION2_RAM, _tb(1, BASTION2_STORAGE), alt=True); r += 1
    _dr(r, None, "Email Service", 1, "Two million mails/month", "",
        None, None, None, None, None, None); r += 1
    if is_prod:
        _dr(r, None, "Back Up - DB", 1,
            "As per banks retention policy. Backup solution can be used.", "",
            None, None, None, None, None, None); r += 1

    # Column widths (col A=stub, B onward)
    ws.column_dimensions["A"].width = 2
    for col_i, w in enumerate([8, 42, 8, 20, 35, 14, 12, 16, 12, 14, 16, 4, 35, 4]):
        ws.column_dimensions[get_column_letter(2+col_i)].width = w


def generate_onprem_excel(
    metrics:      dict,
    distribution: dict,
    customer:     str   = "Bank-Name",
    output_dir:   str   = "reports",
    db_type:      str   = "SQL Server",
    years:        int   = 5,
    filename:     str   = "onprem_openshift_sizing.xlsx",
    cluster_name: str   = None,    # "OpenShift" or "Kubeadm" — overrides auto-detect
    include_dr:   bool  = False,   # Add DR sheet only when user checked the DR checkbox
    env_names:    list  = None,    # e.g. ["Pre-Prod", "SIT", "UAT"] — only add checked envs
    dr_scale:     float = 1.0,     # 0.5 = 50% compute DR, DB full prod; 1.0 = full mirror
) -> str:
    """
    Generates the OpenShift or Kubeadm On-Premise sizing workbook.
    Sheets always included: Data, PROD-1Yr…NYr
    Sheets added conditionally based on user checkbox selections:
      - DR         → only if include_dr=True
      - PRE-PROD   → only if "Pre-Prod" in env_names
      - UAT        → only if "UAT" in env_names
      - SIT        → only if "SIT" in env_names
    No pricing data is written — this is a pure infrastructure sizing file.
    Returns the saved file path.
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    env_names = env_names or []

    metrics_by_year = _onprem_metrics_by_year(metrics, years)

    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    # Data sheet — always present
    _build_onprem_data_sheet(wb, metrics_by_year, customer)

    # PROD sheets — one per year, always present
    # Extract ClickHouse sizing from distribution (if enabled)
    ch_sizing = (distribution or {}).get("clickhouse_nodes", {})
    ch_enabled = ch_sizing and ch_sizing.get("enabled")

    for m in metrics_by_year:
        y = m["year"]
        _build_onprem_env_sheet(
            wb,
            sheet_name=f"PROD-{y}Yr",
            env_label=f"PROD (Year {y})",
            env_data=m,
            customer=customer,
            is_prod=True,
            include_reporting_db=True,
            archival_san_gb=5000,
            db_type=db_type,
            cluster_name=cluster_name,
            ch_sizing=ch_sizing if ch_enabled else None,
        )

    # DR sheet — only when the DR checkbox was checked
    if include_dr:
        import math as _math
        dr_m = dict(metrics_by_year[-1])
        if dr_scale < 1.0:
            # Pilot-light: 1 worker node kept running; infra halved.
            # Nodes are scaled up only when DR is actually activated.
            dr_m["worker_nodes"] = 1
            dr_m["infra_nodes"]  = max(1, _math.ceil(dr_m["infra_nodes"]  * dr_scale))
            # Recompute NFS from scaled worker/infra counts; DB SAN unchanged (san_gb stays same)
            dr_m["nfs_gb"] = (
                dr_m["data_gb"]
                + dr_m["s3_gb"]
                + dr_m["worker_nodes"] * 256
                + dr_m["infra_nodes"]  * 256
                + IMAGE_REGISTRY_GB
            )
            # db_nodes, db_vcpu, db_ram, san_gb — UNCHANGED (full prod DB replication)
        _build_onprem_env_sheet(
            wb,
            sheet_name="DR",
            env_label=f"DR ({int(dr_scale*100)}% Compute — Full Prod DB)",
            env_data=dr_m,
            customer=customer,
            is_prod=True,
            include_reporting_db=True,
            archival_san_gb=5000,
            db_type=db_type,
            cluster_name=cluster_name,
            ch_sizing=ch_sizing if ch_enabled else None,
        )

    # Pre-Prod / SIT / UAT sheets — only when the respective checkbox was checked
    y1 = metrics_by_year[0]
    preprod_m = dict(y1)
    preprod_m["worker_nodes"] = 1
    preprod_m["infra_nodes"]  = 2
    preprod_m["db_nodes"]     = 1
    preprod_m["db_vcpu"]      = y1["db_vcpu"] // max(y1["db_nodes"], 1)
    preprod_m["db_ram"]       = y1["db_ram"]  // max(y1["db_nodes"], 1)
    preprod_m["nfs_gb"]       = y1["data_gb"] + y1["s3_gb"] + 1*256 + 2*256 + IMAGE_REGISTRY_GB
    preprod_m["san_gb"]       = y1["data_gb"]

    if "Pre-Prod" in env_names:
        _build_onprem_env_sheet(
            wb, "PRE-PROD", "PRE-PROD", preprod_m, customer,
            is_prod=False, include_reporting_db=False,
            db_type=db_type, cluster_name=cluster_name,
            ch_sizing=ch_sizing if ch_enabled else None,
        )

    uat_m = dict(y1)
    uat_m["worker_nodes"] = 1
    uat_m["infra_nodes"]  = 2
    uat_m["db_nodes"]     = 1
    uat_m["db_vcpu"]      = y1["db_vcpu"] // max(y1["db_nodes"], 1)
    uat_m["db_ram"]       = y1["db_ram"]  // max(y1["db_nodes"], 1)
    uat_m["data_gb"]      = 500
    uat_m["s3_gb"]        = 500
    uat_m["nfs_gb"]       = 500 + 500 + 1*256 + 2*256 + IMAGE_REGISTRY_GB
    uat_m["san_gb"]       = 500

    if "UAT" in env_names:
        _build_onprem_env_sheet(
            wb, "UAT", "UAT", uat_m, customer,
            is_prod=False, include_reporting_db=False,
            db_type=db_type, cluster_name=cluster_name,
            ch_sizing=ch_sizing if ch_enabled else None,
        )

    if "SIT" in env_names:
        sit_m = dict(uat_m)
        _build_onprem_env_sheet(
            wb, "SIT", "SIT", sit_m, customer,
            is_prod=False, include_reporting_db=False,
            db_type=db_type, cluster_name=cluster_name,
            ch_sizing=ch_sizing if ch_enabled else None,
        )

    out_path = os.path.join(output_dir, filename)
    wb.save(out_path)
    print(f"[excel_exporter] Saved On-Prem sizing ({cluster_name or db_type}): {out_path}")
    return out_path